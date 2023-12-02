from tkinter import filedialog  # 文件访问对话框
from openpyxl import Workbook, load_workbook
import re
import os

# 学校和班级列的索引
school_index = 3
class_index = 4

path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                  defaultextension='.xlsx')
wb = load_workbook(path, read_only=True)
ws = wb.active

# 获取每个学校的所有班级
school_class_dict = {}
for i, row in enumerate(ws.values):
    if i == 0:
        continue
    school_name = row[school_index]
    if school_name not in school_class_dict:
        school_class_dict[school_name] = set()
    school_class_dict[school_name].add(row[class_index])

re_class = re.compile(r'\d+')
school_class_map = {}


def process_class(class_):
    result = re_class.findall(class_)
    if result and len(result[-1]) < 3:
        class_new = int(result[-1])
        if class_new not in class_list_new:
            return class_new
    return class_no


wb_new = Workbook(write_only=True)
ws_new = wb_new.create_sheet('学校班级')
ws_new.append(('学校', '班级', '班级改'))
for school_name, class_set in school_class_dict.items():
    class_list = sorted(class_set)
    class_list_new = []
    class_no = 1

    for class_ in class_list:
        class_new = process_class(class_)
        class_list_new.append(class_new)
        if class_list_new:
            class_no = max(class_list_new) + 1
        ws_new.append((school_name, class_, f'{class_new:0>2d}'))

wb_new.save(f'{os.path.dirname(path)}/学校班级对应.xlsx')
wb_new.close()
wb.close()
