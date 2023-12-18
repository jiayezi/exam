from tkinter import filedialog
from openpyxl import Workbook, load_workbook


# 选科名称的索引
subject_range = (9, 10, 11)
# 考室号的索引
room_index = 6
# 主科
pub_subject = '语文、数学、英语'


def get_student_info():
    file_path = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                           defaultextension='.xlsx')
    if not file_path:
        return
    subject_dict = {pub_subject: [], '政治': [], '历史': [], '地理': [], '物理': [], '化学': [], '生物': []}
    wb = load_workbook(file_path)
    ws = wb.active
    title = next(ws.values)
    # 提取每个科目的学生信息
    for i, row in enumerate(ws.values):
        if i == 0:
            continue
        subject_dict[pub_subject].append(row)
        for j in subject_range:
            subject_name = row[j]
            subject_dict[subject_name].append(row)
    wb.close()

    # 分科目保存
    save_dir = filedialog.askdirectory(title='选择存储文件夹', initialdir='F:/用户目录/桌面/')
    if not save_dir:
        return
    for subject_name,  student_list in subject_dict.items():
        if subject_name == pub_subject:
            continue
        wb = Workbook()
        ws = wb.active
        ws.append(title)
        for row in student_list:
            ws.append(row)
        wb.save(f'{save_dir}/{subject_name}.xlsx')
        wb.close()

    # 计算每个考场的人数
    wb = Workbook()
    ws = wb.active
    ws.append(('科目', '考室', '人数'))
    for subject_name, student_list in subject_dict.items():
        exam_room_dict = {}
        for row in student_list:
            room_name = row[room_index]
            if room_name in exam_room_dict.keys():
                exam_room_dict[room_name] += 1
            else:
                exam_room_dict[room_name] = 1
        counter = 0
        for k, v in exam_room_dict.items():
            ws.append((subject_name, f'第{k}考室', v))
            counter += v
        ws.append((subject_name, '加印', 10))
        ws.append((subject_name, '总计', counter+10))
        ws.append([])
    wb.save(f'{save_dir}/印刷安排.xlsx')


if __name__ == '__main__':
    get_student_info()
