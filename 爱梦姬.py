import os
import shutil
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import messagebox, simpledialog, filedialog  # 消息框，对话框，文件访问对话框
from openpyxl import load_workbook
from PIL import Image
from PIL import UnidentifiedImageError
import fitz  # pymupdf模块，操作PDF文件，提取图片
from win32com import client  # 可操作office文档，转换格式


def initialize():
    """取消冻结文本框，清空文本框"""
    text3.config(state=NORMAL)
    text3.delete(1.0, END)


def word_to_pdf(word_path):
    """word转pdf"""
    file_name = os.path.basename(word_path)
    file_name = file_name[:file_name.rfind('.')]
    if not os.path.exists('tmp'):
        os.mkdir('tmp')
    pdf_path = os.path.join(os.getcwd(), 'tmp', f'{file_name}.pdf')  # 需要使用绝对路径，使用相对路径会出错

    word = client.Dispatch('Word.Application')
    word.Visible = False
    word.DisplayAlerts = False
    doc = word.Documents.Open(word_path)
    doc.SaveAs(pdf_path, FileFormat=17)
    doc.Close()
    word.Quit()
    return pdf_path


def long_png(path, output_name='0'):
    """纵向拼接图片"""
    img_list = []
    for file in os.listdir(path):
        if file.endswith('.png'):
            img_list.append(Image.open(path + os.sep + file))

    width = 0
    height = 0
    for img in img_list:
        # 单幅图像尺寸
        w, h = img.size
        # 获取总高度
        height += h
        # 取最大的宽度作为拼接图的宽度
        width = max(width, w)

    # 创建白色的空白长图
    result = Image.new(mode='RGB', size=(width, height), color=0xffffff)
    # 拼接图片
    height = 0
    for img in img_list:
        w, h = img.size
        # 图片水平居中
        result.paste(img, box=(round(width / 2 - w / 2), height))
        height += h

    # 保存图片
    save_path = filedialog.asksaveasfilename(title='请选择图片存储路径',
                                             initialdir='F:/用户目录/桌面/',
                                             initialfile=output_name,
                                             filetypes=[('PNG', '.png')],
                                             defaultextension='.png')
    if save_path:
        result.save(save_path)
        text3.insert(END, '图片保存成功\n')


def word_to_images():
    """word转图片"""

    word_path = filedialog.askopenfilename(title='请选择Word文档',
                                           filetypes=[('Word文档', '.docx'), ('Word文档', '.doc')],
                                           defaultextension='.docx')
    pdf_path = word_to_pdf(word_path)
    pdf_to_images(pdf_path)


def pdf_to_images(pdf_path=''):
    """pdf转图片"""

    if not os.path.exists('tmp'):
        os.mkdir('tmp')

    if pdf_path == '':
        # 打开PDF文件，生成一个对象
        pdf_path = filedialog.askopenfilename(title='请选择PDF文档',
                                              filetypes=[('PDF', '.pdf')],
                                              defaultextension='.pdf')

    if pdf_path:
        initialize()
        file_name = os.path.basename(pdf_path)
        file_name = file_name[:file_name.rfind('.')]

        pdf = fitz.open(pdf_path)
        for page in pdf:
            pm = page.get_pixmap(dpi=150)
            pm.save(f'tmp/{page.number:0>3d}.png')
        pdf.close()

        text3.insert(END, '已转换成图片\n')
        text3.tag_add('forever', 1.0, END)

        long_png('tmp', file_name)

        # 删除临时文件
        for file in os.listdir('tmp'):
            try:
                os.remove(f'tmp/{file}')
            except PermissionError:
                pass

    over()


def xuanze():
    """根据字母输出对应的字母图片"""
    initialize()
    data = text1.get(1.0, END)
    data = data.strip()
    if data:
        img_path = filedialog.askdirectory(title='请选择答案文件夹', initialdir='F:/用户目录/桌面/')
        if img_path:
            num = 0
            for s in data:
                if s in ('A', 'B', 'C', 'D', 'E', 'F', 'G'):
                    num += 1
                    shutil.copyfile(f'img/{s}.png', f'{img_path}/{num}.png')
            text3.insert(END, f'制作了{num}个答案\n')
    else:
        text3.insert(END, '请先输入答案\n')

    over()


def pinjie():
    """拼接两个文件夹里的名字相同的图片，拼接成功后删除第二个文件夹的图片"""
    img_dir = filedialog.askdirectory(title='请选择题目文件夹', initialdir='F:/用户目录/桌面/')
    if img_dir:
        da_dir = filedialog.askdirectory(title='请选择答案文件夹', initialdir='F:/用户目录/桌面/')
        if da_dir:
            initialize()

            img_list = os.listdir(img_dir)

            for img in img_list:
                if os.path.exists(da_dir + os.sep + img):
                    try:
                        p_ti = Image.open(img_dir + os.sep + img)
                        p_da = Image.open(da_dir + os.sep + img)
                    except UnidentifiedImageError:
                        text3.insert(END, '一个文件不是图片格式，打开失败\n')
                        continue
                    ti_height = p_ti.height
                    new_width = max(p_ti.width, p_da.width)
                    new_height = p_ti.height + p_da.height
                    # 多留10像素，左右两边各留5像素的空白背景
                    result = Image.new(mode='RGB', size=(new_width + 10, new_height), color=(255, 255, 255))
                    result.paste(p_ti, box=(5, 0))
                    result.paste(p_da, box=(5, ti_height))
                    result.save(img_dir + os.sep + img)
                    os.remove(da_dir + os.sep + img)
                else:
                    text3.insert(END, f'{img[:-4]}没有答案\n')
            # 删除空文件夹
            if not os.listdir(da_dir):
                os.rmdir(da_dir)

            text3.insert(END, '拼接完成\n')
    over()


def pic_name():
    """根据分隔符两边的数字确认图片的数量，复制图片并改名"""
    img_list = filedialog.askopenfilenames(title='请选择图片文件',
                                           filetypes=[('PNG', '.png'), ('JPG', '.jpg')],
                                           defaultextension='.png')
    if img_list:
        initialize()
        for path in img_list:
            file_directory = os.path.dirname(path)
            file_name = os.path.basename(path)
            name = file_name[:file_name.rfind('.')]  # 文件名
            fs = ''
            if name[0] in ('A', 'B'):
                fs = name[0]
                name = name[1:]
            extension = file_name[file_name.rfind('.'):]  # 扩展名
            num_list = name.split('-')
            if len(num_list) == 2:
                if num_list[0].isdigit() and num_list[1].isdigit():
                    start = int(num_list[0])
                    end = int(num_list[1])
                    for i in range(start, end):
                        shutil.copyfile(path, f'{file_directory}/{fs}{i}{extension}')
                    os.rename(path, f'{file_directory}/{fs}{end}{extension}')
                else:
                    text3.insert(END, f'文件名 {name} 不含数字，跳过修改 (ー_ー)!!\n')
            else:
                text3.insert(END, f'文件名 {name} 格式不正确，跳过修改 (ー_ー)!!\n')
        text3.insert(END, '修改完成\n')
    over()


def pic_num():
    """把文件夹里的图片名改成Excel里的编号，复制文件夹并改名"""
    file_name = filedialog.askopenfilename(title='请选择Excel文件',
                                           initialdir='F:/用户目录/桌面/',
                                           filetypes=[('Excel', '.xlsx')],
                                           defaultextension='.xlsx')

    if file_name:
        initialize()
        subject = {'01': '语文', '02': '数学', '03': '数学文', '04': '数学理', '05': '英语', '06': '政治', '07': '历史',
                   '08': '地理', '09': '物理', '10': '化学', '11': '生物', '13': '科学', '14': '品德与社会',
                   '15': '道德与法治'}
        wb = load_workbook(file_name)
        ws = wb.worksheets[0]

        while True:
            num = simpledialog.askstring(' ', '请输入科目编号：')
            if num in subject:
                text3.insert(END, f'科目是 {subject[num]}\n')
                text3.tag_add('forever', 1.0, END)
            else:
                subject[num] = simpledialog.askstring(' ', '没有找到科目，请输入科目名称：')

            img_dir = filedialog.askdirectory(title='请选择图片文件夹', initialdir='F:/用户目录/桌面/')

            if img_dir:
                complete = True
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row, 2).value == subject[num]:
                        img_name = ws.cell(row, 3).value
                        abs_img_name = os.path.join(img_dir, f'{img_name}.png')
                        img_id = ws.cell(row, 1).value
                        abs_img_id = os.path.join(img_dir, f'{img_id}.png')
                        if os.path.exists(abs_img_name):
                            shutil.copyfile(abs_img_name, abs_img_id)
                        else:
                            text3.insert(END, f'图片 {img_name}.png 不存在 (ー_ー)!!\n')
                            complete = False

                text3.insert(END, '图片文件名修改完成\n')

                # 复制图片到指定目录
                if complete:
                    img_list = os.listdir(img_dir)
                    if not os.path.exists(f'{img_dir}/03'):
                        os.mkdir(f'{img_dir}/03')
                    for img in img_list:
                        shutil.move(f'{img_dir}/{img}', f'{img_dir}/03')
                    shutil.copytree(f'{img_dir}/03', f'{img_dir}/13')
                    p_dir = img_dir[:img_dir.rfind('/')]
                    if not os.path.exists(f'{p_dir}/{num}'):
                        os.rename(img_dir, f'{p_dir}/{num}')
                    text3.insert(END, '文件复制完成 (＾▽＾) \n')

            else:
                text3.insert(END, '没有选择图片文件夹\n')

            choice = messagebox.askyesno('改名确认', '是否继续改名？')
            if not choice:
                break
        wb.close()
    over()


def kemu_dir(event):
    choice = messagebox.askquestion(title=' ', message='是否分科？')
    if choice == 'yes':
        kemu = ('语文', '数学文', '数学理', '英语', '政治', '历史', '地理', '物理', '化学', '生物')
    else:
        kemu = ('语文', '数学', '英语', '政治', '历史', '地理', '物理', '化学', '生物')
    path = filedialog.askdirectory(title='请选择目录')
    if path:
        try:
            os.mkdir(path + os.sep + '题干')
            for d in kemu:
                os.mkdir(f'{path}/题干/{d}')
                os.mkdir(f'{path}/题干/{d}答案')
        except FileExistsError:
            pass


def over():
    """改变文本颜色，禁用文本框"""
    text3.tag_add('forever', 1.0, END)
    text3.config(state=DISABLED)
    text3.yview_moveto(1)  # 文本更新滚动显示


def close_handle():
    if messagebox.askyesno('退出确认', '确定要退出吗？'):
        root.destroy()


def show_message():
    top = ttk.Toplevel()
    top.title('软件介绍')
    top.geometry('500x250+750+280')  # 窗口大小
    top.maxsize(600, 350)
    top.minsize(350, 180)

    text0 = ttk.Text(top, width=800, height=20, spacing1=10, spacing2=10)
    text0.pack()
    text0.insert(END, '本软件用于处理考试成绩相关的图片，减少工作中的复杂和重复的劳动。\n\n'
                      'Word/PDF转长图：字面意思\n'
                      '制作答案：读取单项选择题答案，按顺序生成答案图片，输出图片到指定文件夹。\n'
                      '拼接图片：打开两个图片文件夹，把文件名相同的图片拼接，拼接成功后删除原始图片。\n'
                      '拆文件名：找出文件名以“-”分隔的图片，复制图片并修改图片文件名，如把A1-3.png改成A1.png，A2.png，A3.png。\n'
                      '添加编号：复制题干图片并把图片名字改成试题结构里的系统编号，然后把图片放进 03 和 13 文件夹，修改上级文件夹为科目编号。\n')

    text0.tag_config('forever', foreground='green', font=('黑体', 12), spacing3=5)
    text0.tag_add('forever', 1.0, END)
    text0.config(state=DISABLED)

    top.mainloop()


def about():
    messagebox.showinfo(title='关于', message='爱梦姬 1.0\n'
                                              'by 李清萍\n'
                                              'QQ 1601235906\n')


root = ttk.Window()
root.title('爱梦姬')
root.geometry('600x400+700+200')  # 窗口大小
root.minsize(500, 340)
root.maxsize(600, 550)
root.bind('<Control-m>', kemu_dir)

menubar = ttk.Menu(root)
help_menu = ttk.Menu(menubar, tearoff=0)
help_menu.add_command(label='介绍', command=show_message)
help_menu.add_command(label='关于', command=about)
menubar.add_cascade(label='帮助', menu=help_menu)

text1 = ttk.Text(root, width=60, height=6, border=-1)
text1.pack(pady=15, side=TOP)

text3 = ttk.Text(root, width=60, height=6, border=-1)
text3.pack(pady=0, side=TOP)
text3.insert(END, '请选择功能\n')
text3.tag_config('forever', foreground="green", font=('黑体', 12), spacing3=8, justify=CENTER)
text3.tag_add('forever', 1.0, END)
text3.config(state=DISABLED)

buttonbar = ttk.Frame(root)
buttonbar.pack(padx=10, pady=25, side=BOTTOM)

btn = ttk.Button(master=buttonbar, text='制作答案', compound=LEFT, command=xuanze)
btn.pack(side=LEFT, ipadx=12, padx=10, pady=5)

btn = ttk.Button(master=buttonbar, text='拼接图片', compound=LEFT, command=pinjie)
btn.pack(side=LEFT, ipadx=12, padx=10, pady=5)

btn = ttk.Button(master=buttonbar, text='拆文件名', compound=LEFT, command=pic_name)
btn.pack(side=LEFT, ipadx=12, padx=10, pady=5)

btn = ttk.Button(master=buttonbar, text='添加编号', compound=LEFT, command=pic_num)
btn.pack(side=LEFT, ipadx=12, padx=10, pady=5)

buttonbar2 = ttk.Frame(root)
buttonbar2.pack(padx=10, pady=0, side=BOTTOM)

btn = ttk.Button(master=buttonbar2, text='Word转长图', compound=LEFT, command=word_to_images)
btn.pack(side=LEFT, ipadx=12, padx=10, pady=5)

btn = ttk.Button(master=buttonbar2, text='PDF转长图', compound=LEFT, command=pdf_to_images)
btn.pack(side=LEFT, ipadx=12, padx=10, pady=5)

root.protocol('WM_DELETE_WINDOW', close_handle)  # 点击关闭按钮，触发事件

root.config(menu=menubar)
root.mainloop()
