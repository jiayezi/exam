"""doc批量转换为docx，提取每一页的word表格数据，保存为excel文件"""
import os
from tkinter import filedialog
import docx
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from win32com import client


subject = ('语文', '数学', '数学文', '数学理', '英语', '政治', '历史', '地理', '物理', '化学', '生物', '文科综合', '理科综合')
border1 = Border(left=Side(border_style='thin', color='000000'),
                 right=Side(border_style='thin', color='000000'),
                 top=Side(border_style='thin', color='000000'),
                 bottom=Side(border_style='thin', color='000000'))

word_path = filedialog.askdirectory(title='请选择word文件夹', initialdir='E:/库/桌面/')
word_path_new = os.path.dirname(word_path)+'/docx'
if not os.path.exists(word_path_new):
    os.mkdir(word_path_new)
excel_path = os.path.dirname(word_path)+'/excel'
if not os.path.exists(excel_path):
    os.mkdir(excel_path)

word = client.Dispatch('Word.Application')
word.Visible = False
word.DisplayAlerts = False
word_list = os.listdir(word_path)
for word_file in word_list:
    doc = word.Documents.Open(f"{word_path}/{word_file}")
    doc.SaveAs(f"{word_path_new}/{word_file}.docx", 12)
    doc.Close()
word.Quit()

# 打开Word文档，提取表格数据
word_list = os.listdir(word_path_new)
for word_file in word_list:
    document = docx.Document(rf"{word_path_new}\{word_file}")
    class_ = word_file.split('_')[0]
    id_number = ''
    name = ''
    end_tag = False
    page_data = []
    for table in document.tables:
        table_data = []
        for row in table.rows:
            row_data = tuple(map(lambda x: x.text, row.cells))
            if row_data[0] == '考号':
                id_number = row_data[1]
                name = row_data[3]
            if row_data[0] == '历史':  # 文科和理科的结束科目不一样，需要分开处理，需要修改这里的科目名称
                end_tag = True
            table_data.append(row_data)
        page_data.append(table_data)

        # 一页的表格提取完之后，就保存为一个工作簿
        if end_tag:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f'{name}的成绩单'

            for a_table in page_data:
                for i, row in enumerate(a_table):
                    ws.append(row)
                    # 对有数据的单元格添加边框
                    for col in range(1, ws.max_column + 1):
                        ws.cell(ws.max_row, col).border = border1
                        ws.cell(ws.max_row, col).alignment = Alignment(horizontal='center', vertical='center')
                    # 对奇数行设置背景颜色
                    if i % 2 == 1:
                        for col in range(1, ws.max_column+1):
                            ws.cell(ws.max_row, col).fill = PatternFill("solid", "EEEEEE")
                    # 判断是否需要合并单元格
                    if row[0] in subject:
                        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=16)
                        ws.cell(ws.max_row, 1).alignment = Alignment(horizontal='center', vertical='center')
                ws.append([])

            ws.merge_cells('A1:P1')
            ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
            # 前两行的字体加粗
            for row in range(1, 3):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).font = Font(bold=True)
            # 取消第二行的背景色
            for col in range(1, ws.max_column + 1):
                ws.cell(2, col).fill = PatternFill(fill_type=None)
            # 保存文件
            if not os.path.exists(f'{excel_path}/{class_}'):
                os.mkdir(f'{excel_path}/{class_}')
            wb.save(f'{excel_path}/{class_}/{name}_{id_number}.xlsx')

            end_tag = False
            page_data = []

