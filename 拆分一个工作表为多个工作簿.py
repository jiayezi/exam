from tkinter import filedialog
import openpyxl


# 指定拆分的列
name_index = 3
# 标题行数
title_rows = 1
# 存储目录
save_path = 'F:/用户目录/桌面/全部学校'

path = filedialog.askopenfilename(title='请选择Excel文件', initialdir='F:/用户目录/桌面/',
                                  filetypes=[('Excel', '.xlsx')], defaultextension='.xlsx')

wb = openpyxl.load_workbook(path, read_only=True)

# 读取数据并分组
title_dict = {}
data_dict = {}
for ws in wb:
    sheetname = ws.title
    title_dict[sheetname] = []
    for j, row in enumerate(ws.values):
        if j < title_rows:
            title_dict[sheetname].append(row)
            continue
        wbname = row[name_index]
        if wbname not in data_dict:
            data_dict[wbname] = {}
        if sheetname not in data_dict[wbname]:
            data_dict[wbname][sheetname] = set()
        data_dict[wbname][sheetname].add(row)

# 生成新工作簿
for wbname, sheet_data in data_dict.items():
    wb_new = openpyxl.Workbook(write_only=True)
    for sheetname, row_data in sheet_data.items():
        ws_new = wb_new.create_sheet(sheetname)
        # 添加标题
        title = title_dict[sheetname]
        for row in title:
            ws_new.append(row)
        # 添加数据
        for row in row_data:
            ws_new.append(row)
        # 合并单元格
        # for rg in ws.merged_cells:
        #     ws_new.merge_cells(str(rg))

    # 保存并关闭工作簿
    wb_new.save(f'{save_path}/{wbname}.xlsx')
    wb_new.close()

wb.close()
