from tkinter import filedialog
from openpyxl import load_workbook, Workbook


def get_count(list, score, i):
    count = 0
    for value in list:
        if value[extra + i] < score:
            count += 1
    return count


extra = 5  # 前5列数据用不上

# 读取Excel文件
file_path = filedialog.askopenfilename(title='请选择Excel文件', initialdir='F:/用户目录/桌面/',
                                       filetypes=[('Excel', '.xlsx')], defaultextension='.xlsx')
if not file_path:
    exit()
wb = load_workbook(file_path)
ws = wb.active
ws_rows = []
for row in ws.values:
    ws_rows.append(list(row))
student_data = ws_rows[1:]
ws_title = ws_rows[0]
subjects = ws_title[extra:]
student_num = len(student_data)

rateT = (1, 0.85, 0.5, 0.15, 0.02, 0)
rateY = ((100, 86), (85, 71), (70, 56), (55, 41), (40, 30))
dict_dj = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E'}

for i, subject in enumerate(subjects):
    student_data.sort(key=lambda x: float(x[extra + i]), reverse=True)

    # 获取原始分等级区间
    rateS = [[float(student_data[0][extra + i])]]
    temp_dj = 0
    rate = (student_num - 1) / student_num
    for j, row in enumerate(student_data):
        current_score = float(row[extra + i])
        previous_score = float(student_data[j - 1][extra + i])
        if j > 0 and current_score != previous_score:
            rate = (student_num - j - 1) / student_num  # 领先率
        # rate = get_count(student_data, row[extra + i], i) / student_num
            for index, value in enumerate(rateT):
                if index == 0:
                    continue
                if rate >= value:
                    if temp_dj != index - 1:
                        temp_dj = index - 1
                        rateS[temp_dj - 1].append(float(student_data[j - 1][extra + i]))
                        rateS.append([float(row[extra + i])])
                    break
    rateS[-1].append(float(student_data[-1][extra + i]))
    print(f'\n{subject}原始分等级区间：{rateS}')

    # 计算赋分成绩
    for row in student_data:
        score = float(row[extra + i])
        xsdj = 0
        for index, dj_score in enumerate(rateS):
            if index == 0:
                continue
            if dj_score[0] >= score >= dj_score[1]:
                xsdj = index
                break
        m = rateS[xsdj][1]
        n = rateS[xsdj][0]
        a = rateY[xsdj][1]
        b = rateY[xsdj][0]
        converts = (b * (score - m) + a * (n - score)) / (n - m)
        converts = round(converts)
        print(f'等级：{xsdj}\tm：{m}\tn：{n}\ta：{a}\tb：{b}\t原始分：{score:.1f}\t转换分：{converts}')
        row.append(converts)
        row.append(dict_dj[xsdj])
    ws_title.append(f'{subject}转换分')
    ws_title.append(f'{subject}等级')

# 写入Excel文件
wb = Workbook()
ws = wb.active
ws.append(ws_title)
for row in student_data:
    ws.append(row)
file_path = filedialog.asksaveasfilename(title='请选择文件存储路径', initialdir='F:/用户目录/桌面/',
                                         initialfile='赋分成绩',
                                         filetypes=[('Excel', '.xlsx')], defaultextension='.xlsx')
if file_path:
    wb.save(file_path)

"""
  * @param m 原始分开始
  * @param n 原始分结束
  * @param a 等级赋值分开始
  * @param b 等级赋值分结束
  * @param score 实考分数
"""
