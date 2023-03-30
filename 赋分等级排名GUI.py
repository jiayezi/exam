"""
图像界面程序，计算赋分成绩和排名
版本：1
"""
from threading import Thread
from tkinter import filedialog, messagebox

import ttkbootstrap as ttk
from openpyxl import Workbook, load_workbook
from ttkbootstrap.constants import *


class Student:
    def __init__(self, row):
        self.row = row


class MyThread(Thread):
    def __init__(self, func, *args):
        super().__init__()
        self.func = func
        self.args = args
        self.setDaemon(True)
        self.start()  # 在这里开始

    def run(self):
        self.func(*self.args)


class App(ttk.Frame):

    def __init__(self, index_frame):
        super().__init__(index_frame, padding=10)
        self.checkbutton_var = []
        self.checkbutton_name = []
        self.selected_subject_name = []
        self.selected_subject_index = []
        self.title = []
        self.student_objs = []
        self.wb = None
        self.ws = None

        self.createUI()

    def createUI(self):
        """创建界面元素"""
        self.grid(row=0, column=0)

        self.btn_frame = ttk.Frame(master=self, padding=(20, 0, 20, 0))
        self.btn_frame.grid(row=0, column=0, padx=80, pady=10, sticky='n')

        ttk.Label(master=self.btn_frame, text='功能', font=('黑体', 12)).grid(row=0, column=0, padx=5, pady=10)
        self.open_btn = ttk.Button(master=self.btn_frame, text='选择文档', command=self.open_file)
        self.open_btn.grid(row=1, column=0, pady=10)
        self.submit_btn = ttk.Button(master=self.btn_frame, text='提交数据',
                                     command=lambda: MyThread(self.extract_data), state=DISABLED)
        self.submit_btn.grid(row=2, column=0, pady=10)
        self.convert_btn = ttk.Button(master=self.btn_frame, text='成绩赋分',
                                      command=lambda: MyThread(self.create_convert_page),
                                      state=DISABLED)
        self.convert_btn.grid(row=3, column=0, pady=10)
        self.rank_btn = ttk.Button(master=self.btn_frame, text='计算排名',
                                   command=lambda: MyThread(self.create_rank_page),
                                   state=DISABLED)
        self.rank_btn.grid(row=4, column=0, pady=10)
        self.save_btn = ttk.Button(master=self.btn_frame, text='保存文件', command=lambda: MyThread(self.save_file),
                                   state=DISABLED)
        self.save_btn.grid(row=5, column=0, pady=10)

    def open_file(self):
        """打开Excel，创建复选框"""
        path = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                          defaultextension='.xlsx')
        if not path:
            return

        self.wb = load_workbook(path, read_only=True)
        self.ws = self.wb.active
        self.submit_btn.config(state=NORMAL)

    def extract_data(self):
        """提取数据"""
        self.open_btn.config(state=DISABLED)
        self.submit_btn.config(state=DISABLED)

        self.title = list(next(self.ws.values))

        # 存为对象
        self.student_objs = []
        for i, row in enumerate(self.ws.values):
            if i == 0:
                continue
            stu_obj = Student(list(row))
            self.student_objs.append(stu_obj)
        self.wb.close()

        messagebox.showinfo(message='提交成功')
        self.open_btn.config(state=NORMAL)
        self.submit_btn.config(state=NORMAL)
        self.convert_btn.config(state=NORMAL)
        self.rank_btn.config(state=NORMAL)
        self.save_btn.config(state=NORMAL)

    def create_convert_page(self):
        def back():
            convert_page.destroy()
            self.grid()

        def convert_score():
            """成绩赋分"""

            def sort_rule(score):
                """定义排序规则"""
                if score is None or score == '':
                    return 0
                else:
                    return float(score)

            # 获取选中的科目的下标和名字
            self.selected_subject_index = []
            self.selected_subject_name = []
            for value in self.checkbutton_var:
                data = value.get()
                if data:
                    self.selected_subject_name.append(data)
                    index = self.title.index(data)
                    self.selected_subject_index.append(int(index))

            # 配置领先率、赋分区间和等级
            rateT = (1, 0.97, 0.9, 0.74, 0.50, 0.26, 0.1, 0.03, 0)
            rateY = ((100, 91), (90, 81), (80, 71), (70, 61), (60, 51), (50, 41), (40, 31), (30, 21))
            dict_dj = {0: 'A', 1: 'B+', 2: 'B', 3: 'C+', 4: 'C', 5: 'D+', 6: 'D', 7: 'E'}

            for sub_index, subject in enumerate(self.selected_subject_name):
                score_index = self.selected_subject_index[sub_index]
                self.student_objs.sort(key=lambda x: sort_rule(x.row[score_index]), reverse=True)

                # 获取得分大于0分的人数、获取大于0分的最小原始分
                student_data_reverse = self.student_objs[::-1]
                student_num = len(student_data_reverse)
                min_score = 0.0
                for row_index, student in enumerate(student_data_reverse):
                    if student.row[score_index] is None or student.row[score_index] == '':
                        continue
                    if float(student.row[score_index]) > 0.0:
                        student_num -= row_index
                        min_score = float(student.row[score_index])
                        break

                # 获取原始分等级区间
                rateS = [[float(self.student_objs[0].row[score_index])]]
                temp_dj = 0
                rate = (student_num - 1) / student_num
                previous_score = -1  # 上个分数，初始值为-1
                for row_index, student in enumerate(self.student_objs):
                    current_score_str = student.row[score_index]
                    if current_score_str is None or current_score_str == '' or float(current_score_str) < 0.001:
                        continue

                    current_score = float(current_score_str)
                    if current_score != previous_score:
                        previous_score = current_score
                        rate = (student_num - row_index - 1) / student_num  # 领先率
                        for v_index, value in enumerate(rateT):
                            if v_index == 0:
                                continue
                            if rate >= value:
                                if temp_dj != v_index - 1:
                                    temp_dj = v_index - 1
                                    rateS[temp_dj - 1].append(
                                        float(self.student_objs[row_index - 1].row[score_index]))
                                    rateS.append([float(student.row[score_index])])
                                break
                # rateS[-1].append(float(student_data[-1][extra + i]))
                rateS[-1].append(min_score)
                # print(f'{subject}原始分等级区间：{rateS}')

                # 计算赋分成绩
                for student in self.student_objs:
                    score_str = student.row[score_index]
                    if score_str is None or score_str == '' or float(score_str) < 0.001:
                        student.row.append('')
                        student.row.append('')
                        continue
                    score = float(score_str)
                    xsdj = 0
                    for index, dj_score in enumerate(rateS):
                        if index == 0:
                            continue
                        if dj_score[0] >= score >= dj_score[1]:
                            xsdj = index
                            break
                    m = rateS[xsdj][1]
                    n = rateS[xsdj][0]
                    a = rateY[xsdj][1]
                    b = rateY[xsdj][0]
                    converts = (b * (score - m) + a * (n - score)) / (n - m)
                    student.row.append(round(converts))
                    student.row.append(dict_dj[xsdj])
                self.title.append(f'{subject}转换分')
                self.title.append(f'{subject}等级')
            messagebox.showinfo(message='赋分完成')

        convert_page = ttk.Frame(master=app, padding=20)
        item_frame = ttk.Frame(master=convert_page, padding=(20, 0, 20, 0))
        item_frame.grid(row=0, column=0)
        btn_frame = ttk.Frame(master=convert_page, padding=(20, 0, 20, 0))
        btn_frame.grid(row=0, column=1)
        ttk.Label(master=item_frame, text='赋分科目', font=('黑体', 12)).grid(row=0, column=0, pady=10)

        possible_subjects = (
            '语文', '数学', '数学文', '数学理', '英语', '外语', '政治', '历史', '地理', '物理', '化学', '生物', '总分')
        self.checkbutton_var = []
        self.checkbutton_name = []
        for i, item in enumerate(self.title):
            if item in possible_subjects:
                self.checkbutton_var.append(ttk.StringVar())
                cb = ttk.Checkbutton(master=item_frame, text=f'{i + 1:0>2d} {item}',
                                     variable=self.checkbutton_var[-1],
                                     onvalue=item, offvalue='')
                cb.grid(row=i + 1, column=0, pady=3, sticky='w')
                self.checkbutton_name.append(item)

        ttk.Button(master=btn_frame, text='计算转换分', command=convert_score).grid(row=0, column=0, pady=10)
        ttk.Button(master=btn_frame, text='返回主页', command=back).grid(row=1, column=0, pady=10)

        self.grid_forget()
        convert_page.grid()

    def create_rank_page(self):
        def back():
            rank_page.destroy()
            self.grid()

        def rank_score():
            """成绩赋分"""

            def sort_rule(score):
                """定义排序规则"""
                if score is None or score == '':
                    return 0
                else:
                    return float(score)

            # 获取选中的科目的下标和名字
            self.selected_subject_index = []
            self.selected_subject_name = []
            for value in self.checkbutton_var:
                data = value.get()
                if data:
                    self.selected_subject_name.append(data)
                    index = self.title.index(data)
                    self.selected_subject_index.append(int(index))

            # 获取选中的排序分组的下标和名字
            self.rank_group_index = []
            self.rank_group_name = []
            for value in self.checkbutton_rank_group_var:
                data = value.get()
                if data:
                    self.rank_group_name.append(data)
                    index = self.title.index(data)
                    self.rank_group_index.append(int(index))

            # 排序的组名
            group_title = ''
            for item in self.rank_group_name:
                group_title += item

            # 给学生设置分组
            group_set = set()
            for index, student in enumerate(self.student_objs):
                group_name = ''
                for g_index in self.rank_group_index:
                    group_name += student.row[g_index]
                group_set.add(group_name)
                student.group = group_name

            # print(f'排序分组:{self.rank_group_name}')
            # print(f'排序科目:{self.selected_subject_name}')

            for sub_index, subject in enumerate(self.selected_subject_name):
                group_list = list(group_set)
                for group in group_list:
                    student_objs_new = list(filter(lambda x: x.group == group, self.student_objs))
                    score_index = self.selected_subject_index[sub_index]
                    student_objs_new.sort(key=lambda x: sort_rule(x.row[score_index]), reverse=True)

                    prev = -1  # 上个分数，初始值为-1
                    rank = 0  # 当前排名
                    for s_index, student in enumerate(student_objs_new):
                        score_str = student.row[score_index]
                        if score_str is None or score_str == '' or float(score_str) < 0.001:
                            student.row.append('')
                            continue
                        score = float(score_str)
                        # 如果分数不一样，排名就是索引值+1，如果分数一样，排名不变
                        if score != prev:
                            rank = s_index + 1
                            prev = score
                        student.row.append(rank)
                self.title.append(f'{subject}{group_title}排名')

            messagebox.showinfo(message='排名计算完成')

        rank_page = ttk.Frame(master=app, padding=20)
        item_frame = ttk.Frame(master=rank_page, padding=(20, 0, 20, 0))
        item_frame.grid(row=0, column=0, sticky='n')
        item_frame2 = ttk.Frame(master=rank_page, padding=(20, 0, 20, 0))
        item_frame2.grid(row=0, column=1, sticky='n')
        btn_frame = ttk.Frame(master=rank_page, padding=(20, 0, 20, 0))
        btn_frame.grid(row=0, column=2)
        ttk.Label(master=item_frame, text='排名科目', font=('黑体', 12)).grid(row=0, column=0, pady=10, sticky='w')
        ttk.Label(master=item_frame2, text='排名分组', font=('黑体', 12)).grid(row=0, column=1, pady=10, sticky='w')

        possible_subjects = ['语文', '数学', '数学文', '数学理', '英语', '外语', '政治', '历史', '地理', '物理', '化学',
                             '生物', '总分']
        possible_subjects_convert = [f'{item}转换分' for item in possible_subjects]
        possible_subjects_new = possible_subjects + possible_subjects_convert

        self.checkbutton_var = []
        self.checkbutton_name = []
        for i, item in enumerate(self.title):
            if item in possible_subjects_new:
                self.checkbutton_var.append(ttk.StringVar())
                cb = ttk.Checkbutton(master=item_frame, text=f'{i + 1:0>2d} {item}',
                                     variable=self.checkbutton_var[-1],
                                     onvalue=item, offvalue='')
                cb.grid(row=i + 1, column=0, pady=3, sticky='w')
                self.checkbutton_name.append(item)

        self.checkbutton_rank_group_var = []
        self.checkbutton_rank_group_name = []
        for i, item in enumerate(self.title):
            if item not in possible_subjects_new:
                self.checkbutton_rank_group_var.append(ttk.StringVar())
                cb = ttk.Checkbutton(master=item_frame2, text=f'{i + 1:0>2d} {item}',
                                     variable=self.checkbutton_rank_group_var[-1],
                                     onvalue=item, offvalue='')
                cb.grid(row=i + 1, column=1, pady=3, sticky='w')
                self.checkbutton_rank_group_name.append(item)

        ttk.Button(master=btn_frame, text='计算排名', command=rank_score).grid(row=0, column=0, pady=10)
        ttk.Button(master=btn_frame, text='返回主页', command=back).grid(row=1, column=0, pady=10)

        self.grid_forget()
        rank_page.grid()

    def save_file(self):
        path = filedialog.asksaveasfilename(title='请选择文件存储位置',
                                            initialdir='F:/用户目录/桌面/',
                                            initialfile='计算结果',
                                            filetypes=[('Excel', '.xlsx')],
                                            defaultextension='.xlsx')
        if not path:
            return

        wb = Workbook(write_only=True)
        ws = wb.create_sheet()

        ws.append(self.title)
        for student in self.student_objs:
            ws.append(student.row)

        wb.save(path)
        wb.close()
        messagebox.showinfo(message='文件保存成功')


def close_handle():
    if messagebox.askyesno(title='退出确认', message='确定要退出吗？'):
        app.destroy()


if __name__ == "__main__":
    app = ttk.Window(title="成绩计算程序")
    # app.geometry(f'650x380')  # 窗口大小
    # app.minsize(333, 333)
    app.iconbitmap('green_apple.ico')
    App(app)
    app.protocol('WM_DELETE_WINDOW', close_handle)  # 启用协议处理机制，点击关闭时按钮，触发事件
    app.place_window_center()
    app.mainloop()
