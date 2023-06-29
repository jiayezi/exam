import os
import openpyxl
from openpyxl.styles import Border, Side, Font, colors, Alignment
from tkinter import filedialog

save_path = 'E:/库/桌面/全部学校'

# 定义格式
font_Calibri_10 = Font(name='Calibri', size=10)
border_thin = Border(left=Side(border_style='thin', color='000000'),
                     right=Side(border_style='thin', color='000000'),
                     top=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000'))

path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                  defaultextension='.xlsx')

wb = openpyxl.load_workbook(path)

# 获取全部学校名字，填充学校名字
schools = []
for index, sheet in enumerate(wb.sheetnames):
    ws = wb[sheet]
    for row in range(4, ws.max_row + 1):
        if ws.cell(row, 2).value == '' or ws.cell(row, 2).value is None:
            ws.cell(row, 2, ws.cell(row-1, 2).value)
        else:
            # 只获取第一个工作表里的全部学校名字
            if index == 0:
                schools.append(ws.cell(row, 2).value)
print(f'一共 {len(schools)} 个学校')

# 提取每个学校的数据
for school in schools:
    wb_new = openpyxl.Workbook()

    # 根据原始工作簿的工作表创建新工作簿的工作表，然后添加数据
    for index, sheet in enumerate(wb.sheetnames):
        ws = wb[sheet]
        if index == 0:
            ws_new = wb_new.active
            ws_new.title = sheet
        else:
            ws_new = wb_new.create_sheet(sheet)

        # 添加前3行
        for i, row in enumerate(ws.values):
            if i == 3:
                break
            ws_new.append(row)

        # 添加学校数据
        for row in ws.values:
            if row[1] == school:
                ws_new.append(row)

        # 添加边框、对齐方式、字体
        data_range = ws_new[ws_new.dimensions]
        for row in data_range:
            for cell in row:
                cell.border = border_thin
                cell.font = font_Calibri_10
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 合并单元格
        for rg in ws.merged_cells:
            ws_new.merge_cells(str(rg))

    # 保存并关闭工作簿
    save_school_path = f'{save_path}/{school}'
    if not os.path.exists(save_school_path):
        os.makedirs(save_school_path)
    wb_new.save(f'{save_school_path}/{school}_班级小题均分.xlsx')
    wb_new.close()

wb.close()
