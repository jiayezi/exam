import os
import openpyxl
from openpyxl.styles import Border, Side, Font, colors, Alignment
from tkinter import filedialog

save_path = 'E:/库/桌面/全部学校'

font_Calibri_10 = Font(name='Calibri', size=10)
border_thin = Border(left=Side(border_style='thin', color='000000'),
                     right=Side(border_style='thin', color='000000'),
                     top=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000'))

path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                  defaultextension='.xlsx')

wb = openpyxl.load_workbook(path)
ws = wb.active

# 获取全部学校名字，填充学校名字
schools = []
for row in range(4, ws.max_row + 1):
    if ws.cell(row, 2).value == '' or ws.cell(row, 2).value is None:
        ws.cell(row, 2, ws.cell(row - 1, 2).value)
    else:
        schools.append(ws.cell(row, 2).value)
print(f'一共 {len(schools)} 个学校')

for school in schools:
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = '班级科目均分'

    # 添加前3行
    for i, row in enumerate(ws.values):
        if i == 3:
            break
        ws2.append(row)

    # 添加学校数据
    for row in ws.values:
        if row[1] == school:
            ws2.append(row)

    # 添加边框、对齐方式、字体
    data_range = ws2[ws2.dimensions]
    for row in data_range:
        for cell in row:
            cell.border = border_thin
            cell.font = font_Calibri_10
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 合并单元格
    for rg in ws.merged_cells:
        ws2.merge_cells(str(rg))

    # 保存并关闭
    save_school_path = f'{save_path}/{school}'
    if not os.path.exists(save_school_path):
        os.makedirs(save_school_path)
    wb2.save(f'{save_school_path}/{school}_班级科目均分.xlsx')
    wb2.close()

wb.close()
