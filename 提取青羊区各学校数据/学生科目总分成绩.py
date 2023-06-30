import os
import openpyxl
from openpyxl.styles import Border, Side, Font, colors, Alignment
from tkinter import filedialog


font_Calibri_10 = Font(name='Calibri', size=10)
border_thin = Border(left=Side(border_style='thin', color='000000'),
                     right=Side(border_style='thin', color='000000'),
                     top=Side(border_style='thin', color='000000'),
                     bottom=Side(border_style='thin', color='000000'))

path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                  defaultextension='.xlsx')
save_path = os.path.dirname(path)+'/全部学校'
wb = openpyxl.load_workbook(path)
ws = wb.active

# 获取全部学校名字
schools = []
for row in range(2, ws.max_row + 1):
    if ws.cell(row, 2).value not in schools:
        schools.append(ws.cell(row, 2).value)
print(f'一共 {len(schools)} 个学校')

title = next(ws.values)   # ws.values是生成器

# 处理每个学校的数据
for school in schools:
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = '学生科目总成绩'

    # 添加标题
    ws2.append(title)

    # 添加学校部分
    for row in ws.values:
        if row[1] == school:
            ws2.append(row)

    # 添加边框、对齐方式、字体
    data_range = ws2[ws2.dimensions]
    for row in data_range:
        for cell in row:
            cell.border = border_thin
            cell.font = font_Calibri_10
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 保存并关闭
    save_school_path = f'{save_path}/{school}'
    if not os.path.exists(save_school_path):
        os.mkdir(save_school_path)
    wb2.save(f'{save_school_path}/{school}_学生科目总分成绩.xlsx')
    wb2.close()

wb.close()

