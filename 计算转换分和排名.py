"""
图像界面程序，计算赋分成绩和排名
版本：1
"""
import os
from threading import Thread
from tkinter import filedialog, messagebox

import ttkbootstrap as ttk
from openpyxl import Workbook, load_workbook
from ttkbootstrap.constants import *


class Student:
    def __init__(self, row):
        self.row = row
        self.group = ''


class MyThread(Thread):
    def __init__(self, func, *args):
        super().__init__()
        self.func = func
        self.args = args
        self.setDaemon(True)
        self.start()  # 在这里开始

    def run(self):
        self.func(*self.args)


def sort_rule(score):
    """定义排序规则"""
    if score is None or score == '':
        return 0
    else:
        return float(score)


def select_all1(select_all_var, checkbutton_var, checkbutton_name):
    """全选和取消全选"""
    if select_all_var.get():
        for index, var in enumerate(checkbutton_var):
            name = checkbutton_name[index]
            var.set(name)
    else:
        for var in checkbutton_var:
            var.set('')


def close_handle():
    if messagebox.askyesno(message='确定要退出吗？'):
        app.destroy()


class App(ttk.Frame):

    def __init__(self, master):
        super().__init__(master, padding=10)
        self.title = []
        self.student_objs = []
        self.createUI()

    def createUI(self):
        """创建界面元素"""
        self.grid(row=0, column=0)

        self.btn_frame = ttk.Frame(master=self, padding=(20, 0, 20, 0))
        self.btn_frame.grid(row=0, column=0, padx=80, pady=10, sticky='n')

        ttk.Label(master=self.btn_frame, text='选择功能', font=('黑体', 12)).grid(row=0, column=0, pady=10)
        self.open_btn = ttk.Button(master=self.btn_frame, text='打开文档', command=self.open_file)
        self.open_btn.grid(row=1, column=0, pady=10)
        self.convert_btn = ttk.Button(master=self.btn_frame, text='成绩赋分',
                                      command=lambda: MyThread(self.create_convert_page),
                                      state=DISABLED)
        self.convert_btn.grid(row=2, column=0, pady=10)
        self.rank_btn = ttk.Button(master=self.btn_frame, text='计算排名',
                                   command=lambda: MyThread(self.create_rank_page),
                                   state=DISABLED)
        self.rank_btn.grid(row=3, column=0, pady=10)
        self.save_btn = ttk.Button(master=self.btn_frame, text='保存文档', command=lambda: MyThread(self.save_file),
                                   state=DISABLED)
        self.save_btn.grid(row=4, column=0, pady=10)
        self.info_text = ttk.StringVar()
        ttk.Label(master=self, textvariable=self.info_text, foreground='#666666', font=('黑体', 12)).grid(row=1, column=0, pady=10)

    def btn_freeze(self):
        self.open_btn.config(state=DISABLED)
        self.convert_btn.config(state=DISABLED)
        self.rank_btn.config(state=DISABLED)
        self.save_btn.config(state=DISABLED)

    def btn_unfreeze(self):
        self.open_btn.config(state=NORMAL)
        self.convert_btn.config(state=NORMAL)
        self.rank_btn.config(state=NORMAL)
        self.save_btn.config(state=NORMAL)

    def convert_template_level(self):
        """维护赋分模板的顶层窗口"""
        def save_template(file_name, et_obj):
            """将所有输入框的值保存到文件"""
            # 验证数字
            total = 0
            for row_obj in et_obj:
                for col, et in enumerate(row_obj):
                    if col > 0 and not check(et):
                        messagebox.showinfo(message='存在不是数字的值，保存失败')
                        return
                    if col == 3:
                        total += float(et.get())
            if total != 100:
                messagebox.showinfo(message='占比之和不等于100，保存失败')
                return

            with open(f'conf/{file_name}', 'wt', encoding='utf8') as f:
                text = ''
                for row_obj in et_obj:
                    values = [et.get() for et in row_obj]
                    text += '\t'.join(values)+'\n'
                f.write(text[:-1])
            messagebox.showinfo(message='保存成功')

        def del_template(file_name):
            """删除模板文件"""
            if messagebox.askyesno(message='确定删除？'):
                os.remove(f'conf/{file_name}')
                top.destroy()
                messagebox.showinfo(message='模板已删除')
                # 重启顶层窗口，达到刷新页面的目的
                self.convert_template_level()

        def add_et_row(et_frame, et_obj):
            """在末尾添加一行输入框"""
            row_obj = []
            for col in range(4):
                if col > 0:
                    et = ttk.Entry(et_frame, width=3, validate='focusout')
                    et.config(validatecommand=lambda e=et: check(e))
                else:
                    et = ttk.Entry(et_frame, width=3)
                et.grid(row=len(et_obj)+1, column=col, padx=10, pady=5)
                row_obj.append(et)
            et_obj.append(row_obj)

        def del_et_row(et_obj):
            """删除最后一行输入框"""
            if len(et_obj) > 1:
                for et in et_obj[-1]:
                    et.destroy()
                et_obj.pop()

        def check(et):
            value = et.get()
            try:
                num = float(value)
                if num > 0:
                    return True
                else:
                    messagebox.showwarning(message='请输入大于0的数字！')
                    return False
            except ValueError:
                messagebox.showwarning(message='请输入数字！')
                return False

        def go_to_top():
            """重启顶层窗口，达到刷新页面的目的"""
            top.destroy()
            self.convert_template_level()

        def close_top_level():
            """关闭顶层窗口并恢复按钮状态"""
            top.destroy()
            self.template_btn.config(state=NORMAL)

        def modify_template(template_name):
            """读取指定文件，创建输入控件，把文件内容添加到输入框"""
            modify_frame = ttk.Frame(master=top, padding=(20, 10, 20, 20))
            et_frame = ttk.Frame(master=modify_frame)
            et_frame.grid(row=0, column=0, sticky='n')
            btn_frame = ttk.Frame(master=modify_frame, padding=(10, 0, 0, 0))
            btn_frame.grid(row=0, column=1)

            ttk.Label(master=et_frame, text='等级', font=('黑体', 10)).grid(row=0, column=0, pady=10)
            ttk.Label(master=et_frame, text='高分', font=('黑体', 10)).grid(row=0, column=1, pady=10)
            ttk.Label(master=et_frame, text='低分', font=('黑体', 10)).grid(row=0, column=2, pady=10)
            ttk.Label(master=et_frame, text='占比(%)', font=('黑体', 10)).grid(row=0, column=3, pady=10)

            # 读取文件，创建输入框
            with open(f'conf/{template_name}', 'rt', encoding='utf8') as f:
                data = f.read()
            data_rows = data.split('\n')
            et_obj = []
            for row_index, row in enumerate(data_rows):
                row_obj = []
                values = row.split('\t')
                for col_index, value in enumerate(values):
                    # 除了第一列，其他列都要验证输入的内容
                    if col_index > 0:
                        et = ttk.Entry(et_frame, width=3, validate='focusout')
                        et.config(validatecommand=lambda e=et: check(e))
                    else:
                        et = ttk.Entry(et_frame, width=3)
                    et.grid(row=row_index+1, column=col_index, padx=10, pady=5)
                    et.insert(END, value)
                    row_obj.append(et)
                et_obj.append(row_obj)

            ttk.Label(master=et_frame, text='名称:', font=('黑体', 10)).grid(row=20, column=0, pady=10)
            file_name_et = ttk.Entry(et_frame, width=19)
            file_name_et.grid(row=20, column=1, columnspan=3, pady=10)
            file_name_et.insert('end', template_name)

            ttk.Button(master=btn_frame, text='新增一级', command=lambda: add_et_row(et_frame, et_obj)).grid(padx=5, pady=5)
            ttk.Button(master=btn_frame, text='删除一级', command=lambda: del_et_row(et_obj)).grid(padx=5, pady=5)
            ttk.Button(master=btn_frame, text='保存模板', command=lambda: save_template(file_name_et.get(), et_obj)).grid(padx=5, pady=5)
            ttk.Button(master=btn_frame, text='返回上页', command=go_to_top).grid(padx=5, pady=5)

            list_frame.grid_forget()
            modify_frame.grid(sticky='n')

        self.template_btn.config(state=DISABLED)

        top = ttk.Toplevel()
        top.title('赋分模板维护')
        top.iconbitmap('green_apple.ico')
        top.protocol("WM_DELETE_WINDOW", close_top_level)
        top.place_window_center()

        new_template = '新模板'
        list_frame = ttk.Frame(master=top, padding=(20, 10, 20, 20))
        list_frame.grid(sticky='n')
        ttk.Label(master=list_frame, text='模板列表', font=('黑体', 12)).grid(row=0, column=0, pady=10)
        ttk.Button(master=list_frame, text='添加', command=lambda: modify_template(new_template)).grid(row=0, column=1, padx=5, pady=10)
        ttk.Button(master=list_frame, text='关闭', command=close_top_level).grid(row=0, column=2, padx=5, pady=10)

        # 读取指定目录里的所有文件，创建标签和按钮
        template_list = os.listdir('conf')
        for i, template in enumerate(template_list):
            if template == new_template:
                continue
            ttk.Label(master=list_frame, text=template, font=('黑体', 12), foreground='#336644').grid(row=i+1, column=0, padx=5, pady=5)
            ttk.Button(master=list_frame, text='修改', command=lambda t=template: modify_template(t)).grid(row=i+1, column=1, padx=5, pady=5)
            ttk.Button(master=list_frame, text='删除', command=lambda t=template: del_template(t)).grid(row=i + 1, column=2, padx=5,  pady=5)

        top.mainloop()

    def open_file(self):
        """打开Excel，读取数据"""
        path = filedialog.askopenfilename(title='选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                          defaultextension='.xlsx')
        if not path:
            return

        self.open_btn.config(state=DISABLED)
        self.info_text.set('正在读取数据')

        wb = load_workbook(path, read_only=True)
        ws = wb.active
        self.title = list(next(ws.values))

        # 存为对象
        self.student_objs = []
        for i, row in enumerate(ws.values):
            if i == 0:
                continue
            stu_obj = Student(list(row))
            self.student_objs.append(stu_obj)
        wb.close()

        self.info_text.set('读取完成')
        self.btn_unfreeze()

    def create_convert_page(self):
        """创建转换分页面"""
        def back():
            convert_page.destroy()
            self.grid()

        def convert_score():
            """计算转换分"""
            convert_btn.config(state=DISABLED)
            self.btn_freeze()
            self.info_text.set('正在计算转换分')

            # 获取选中的科目的下标和名字
            selected_subject_index = []
            selected_subject_name = []
            for value in checkbutton_var:
                data = value.get()
                if data:
                    selected_subject_name.append(data)
                    index = self.title.index(data)
                    selected_subject_index.append(index)

            # 加载配置文件，读取领先率、赋分区间和等级
            rate_exceed = []
            rate_dist = []
            grade_dict = {}
            with open(f'conf/{cb.get()}', 'rt', encoding='utf8') as f:
                data = f.read()
            row_list = data.split('\n')
            rate_sum = 0
            for index, row in enumerate(row_list):
                value_list = row.split('\t')
                grade_dict[index] = value_list[0]
                rg = (float(value_list[1]), float(value_list[2]))
                rate_dist.append(rg)
                rate_sum += float(value_list[3])
                value = (100 - rate_sum) / 100.0
                rate_exceed.append(value)

            # 手动配置领先率、赋分区间和等级
            # rate_exceed = (0.97, 0.9, 0.74, 0.50, 0.26, 0.1, 0.03, 0)
            # rate_dist = ((100, 91), (90, 81), (80, 71), (70, 61), (60, 51), (50, 41), (40, 31), (30, 21))
            # grade_dict = {0: 'A', 1: 'B+', 2: 'B', 3: 'C+', 4: 'C', 5: 'D+', 6: 'D', 7: 'E'}

            for sub_index, subject in enumerate(selected_subject_name):
                score_index = selected_subject_index[sub_index]
                self.student_objs.sort(key=lambda x: sort_rule(x.row[score_index]), reverse=True)

                # 获取得分大于0分的人数、获取大于0分的最小原始分
                student_data_reverse = self.student_objs[::-1]
                student_num = len(student_data_reverse)
                min_score = 0.0
                for row_index, student in enumerate(student_data_reverse):
                    if student.row[score_index] is None or student.row[score_index] == '':
                        continue
                    if float(student.row[score_index]) > 0.0:
                        student_num -= row_index
                        min_score = float(student.row[score_index])
                        break

                # 获取原始分等级区间
                rate_src = [[float(self.student_objs[0].row[score_index])]]
                rate = (student_num - 1) / student_num
                previous_score = -1  # 上个分数，初始值为-1
                temp_dj = 0  # 初始等级和索引
                for row_index, student in enumerate(self.student_objs):
                    current_score_str = student.row[score_index]
                    if current_score_str is None or current_score_str == '' or float(current_score_str) < 0.001:
                        continue

                    current_score = float(current_score_str)
                    if current_score != previous_score:
                        previous_score = current_score
                        rate = (student_num - row_index - 1) / student_num  # 领先率
                        for e_index, value in enumerate(rate_exceed):
                            # 如果这个学生的领先率大于rate_exceed里的某一个值，并且temp_dj与上次不同，就把当前分数添加到rate_src里面，然后结束内层循环。如果这个学生的领先率大于rate_exceed里的某一个值，并且temp_dj与上次相同，也要结束内层循环，所以break不能写在if内部。
                            if rate >= value:
                                if temp_dj != e_index:
                                    temp_dj = e_index
                                    rate_src[temp_dj - 1].append(
                                        float(self.student_objs[row_index - 1].row[score_index]))
                                    rate_src.append([float(student.row[score_index])])
                                break
                # rate_src[-1].append(float(student_data[-1][extra + i]))
                rate_src[-1].append(min_score)
                # print(f'{subject}原始分等级区间：{rate_src}')

                # 计算赋分成绩
                for student in self.student_objs:
                    score_str = student.row[score_index]
                    if score_str is None or score_str == '' or float(score_str) < 0.001:
                        student.row.append('')
                        student.row.append('')
                        continue
                    score = float(score_str)
                    xsdj = 0
                    for index, dj_score in enumerate(rate_src):
                        if dj_score[0] >= score >= dj_score[1]:
                            xsdj = index
                            break
                    m = rate_src[xsdj][1]
                    n = rate_src[xsdj][0]
                    a = rate_dist[xsdj][1]
                    b = rate_dist[xsdj][0]
                    converts = (b * (score - m) + a * (n - score)) / (n - m)
                    student.row.append(round(converts))
                    student.row.append(grade_dict[xsdj])
                self.title.append(f'{subject}转换分')
                self.title.append(f'{subject}等级')

            self.info_text.set('转换分计算完成')
            convert_btn.config(state=NORMAL)
            self.btn_unfreeze()

        convert_page = ttk.Frame(master=app, padding=(20, 10, 20, 20))
        item_frame = ttk.Frame(master=convert_page)
        item_frame.grid(row=0, column=0, padx=10, sticky='n')
        btn_frame = ttk.Frame(master=convert_page)
        btn_frame.grid(row=0, column=1, padx=10, sticky='s')
        ttk.Label(master=item_frame, text='赋分科目', font=('黑体', 12)).grid(row=0, column=0, pady=10)

        # 创建复选框
        select_all_var = ttk.StringVar()
        cb = ttk.Checkbutton(master=item_frame, text='全选', variable=select_all_var,
                             onvalue='全选', offvalue='', command=lambda: select_all1(select_all_var, checkbutton_var, checkbutton_name))
        cb.grid(row=1, column=0, pady=5, sticky='w')

        possible_subjects = (
            '语文', '数学', '数学文', '数学理', '英语', '外语', '政治', '历史', '地理', '物理', '化学', '生物', '总分')
        checkbutton_var = []
        checkbutton_name = []
        for i, item in enumerate(self.title):
            if item in possible_subjects:
                checkbutton_var.append(ttk.StringVar())
                cb = ttk.Checkbutton(master=item_frame, text=f'{i + 1:0>2d} {item}',
                                     variable=checkbutton_var[-1],
                                     onvalue=item, offvalue='')
                cb.grid(row=i + 2, column=0, pady=3, sticky='w')
                checkbutton_name.append(item)

        # 创建按钮和标签
        template_files = os.listdir('conf')
        template_files.remove('新模板')
        cb = ttk.Combobox(master=btn_frame, values=template_files, width=14)
        cb.grid(row=0, column=0, pady=10)
        cb.current(0)
        self.template_btn = ttk.Button(master=btn_frame, text='模板维护', command=self.convert_template_level)
        self.template_btn.grid(row=1, column=0, pady=5)
        convert_btn = ttk.Button(master=btn_frame, text='成绩赋分', command=convert_score)
        convert_btn.grid(row=2, column=0, pady=5)
        ttk.Button(master=btn_frame, text='返回主页', command=back).grid(row=3, column=0, pady=5)
        ttk.Label(master=btn_frame, textvariable=self.info_text, foreground='#666666', font=('黑体', 12)).grid(row=4, column=0, pady=5)

        # 隐藏主页，显示转换分页面
        self.grid_forget()
        convert_page.grid()

    def create_rank_page(self):
        """创建计算排名的页面"""
        def back():
            rank_page.destroy()
            self.grid()

        def rank_score():
            """计算排名"""
            rank_btn.config(state=DISABLED)
            self.btn_freeze()
            self.info_text.set('正在计算排名')

            # 获取选中的科目的下标和名字
            selected_subject_index = []
            selected_subject_name = []
            for value in checkbutton_var:
                data = value.get()
                if data:
                    selected_subject_name.append(data)
                    index = self.title.index(data)
                    selected_subject_index.append(index)

            # 获取选中的排序分组的下标和名字
            rank_group_index = []
            rank_group_name = []
            for value in checkbutton_rank_group_var:
                data = value.get()
                if data:
                    rank_group_name.append(data)
                    index = self.title.index(data)
                    rank_group_index.append(index)

            # 排序的组名
            group_title = ''.join(rank_group_name)

            # 给学生设置分组，获取所有分组
            group_set = set()
            for student in self.student_objs:
                group_name = ''
                for g_index in rank_group_index:
                    group_name += student.row[g_index]
                group_set.add(group_name)
                student.group = group_name

            # 开始计算
            group_list = list(group_set)
            for sub_index, subject in enumerate(selected_subject_name):
                for group in group_list:
                    student_objs_new = list(filter(lambda x: x.group == group, self.student_objs))
                    score_index = selected_subject_index[sub_index]
                    student_objs_new.sort(key=lambda x: sort_rule(x.row[score_index]), reverse=True)

                    prev = -1  # 上个分数，初始值为-1
                    rank = 0  # 当前排名
                    for s_index, student in enumerate(student_objs_new):
                        score_str = student.row[score_index]
                        if score_str is None or score_str == '' or float(score_str) < 0.001:
                            student.row.append('')
                            continue
                        score = float(score_str)
                        # 如果分数不一样，排名就是索引值+1，如果分数一样，排名不变
                        if score != prev:
                            rank = s_index + 1
                            prev = score
                        student.row.append(rank)
                self.title.append(f'{subject}{group_title}排名')

            self.info_text.set('排名计算完成')
            rank_btn.config(state=NORMAL)
            self.btn_unfreeze()

        rank_page = ttk.Frame(master=app, padding=(20, 10, 20, 20))
        subject_item_frame = ttk.Frame(master=rank_page)
        subject_item_frame.grid(row=0, column=0, padx=10, sticky='n')
        group_item_frame = ttk.Frame(master=rank_page)
        group_item_frame.grid(row=0, column=1, padx=10, sticky='n')
        btn_frame = ttk.Frame(master=rank_page)
        btn_frame.grid(row=0, column=2, padx=10, sticky='s')
        ttk.Label(master=subject_item_frame, text='排名科目', font=('黑体', 12)).grid(row=0, column=0, pady=10,
                                                                                      sticky='w')
        ttk.Label(master=group_item_frame, text='排名分组', font=('黑体', 12)).grid(row=0, column=0, pady=10,
                                                                                    sticky='w')
        # 创建科目复选框
        select_all_var1 = ttk.StringVar()
        cb = ttk.Checkbutton(master=subject_item_frame, text='全选', variable=select_all_var1,
                             onvalue='全选', offvalue='', command=lambda: select_all1(select_all_var1, checkbutton_var, checkbutton_name))
        cb.grid(row=1, column=0, pady=5, sticky='w')

        possible_subjects = ['语文', '数学', '数学文', '数学理', '英语', '外语', '政治', '历史', '地理', '物理', '化学',
                             '生物', '总分']
        possible_subjects_convert = [f'{item}转换分' for item in possible_subjects]
        possible_subjects_new = possible_subjects + possible_subjects_convert
        checkbutton_var = []
        checkbutton_name = []
        for i, item in enumerate(self.title):
            if item in possible_subjects_new:
                checkbutton_var.append(ttk.StringVar())
                cb = ttk.Checkbutton(master=subject_item_frame, text=f'{i + 1:0>2d} {item}',
                                     variable=checkbutton_var[-1],
                                     onvalue=item, offvalue='')
                cb.grid(row=i + 2, column=0, pady=3, sticky='w')
                checkbutton_name.append(item)

        # 创建分组复选框
        select_all_var2 = ttk.StringVar()
        cb = ttk.Checkbutton(master=group_item_frame, text='全选', variable=select_all_var2,
                             onvalue='全选', offvalue='', command=lambda: select_all1(select_all_var2, checkbutton_rank_group_var, checkbutton_rank_group_name))
        cb.grid(row=1, column=0, pady=5, sticky='w')

        checkbutton_rank_group_var = []
        checkbutton_rank_group_name = []
        for i, item in enumerate(self.title):
            if item not in possible_subjects_new:
                checkbutton_rank_group_var.append(ttk.StringVar())
                cb = ttk.Checkbutton(master=group_item_frame, text=f'{i + 1:0>2d} {item}',
                                     variable=checkbutton_rank_group_var[-1],
                                     onvalue=item, offvalue='')
                cb.grid(row=i + 2, column=0, pady=3, sticky='w')
                checkbutton_rank_group_name.append(item)

        # 创建按钮
        rank_btn = ttk.Button(master=btn_frame, text='计算排名', command=rank_score)
        rank_btn.grid(row=0, column=0, pady=5)
        ttk.Button(master=btn_frame, text='返回主页', command=back).grid(row=1, column=0, pady=5)
        ttk.Label(master=btn_frame, textvariable=self.info_text, foreground='#666666', font=('黑体', 12)).grid(row=2, column=0, pady=10)

        self.grid_forget()
        rank_page.grid()

    def save_file(self):
        """生成Excel文件"""
        path = filedialog.asksaveasfilename(title='请选择文件存储位置',
                                            initialdir='F:/用户目录/桌面/',
                                            initialfile='计算结果',
                                            filetypes=[('Excel', '.xlsx')],
                                            defaultextension='.xlsx')
        if not path:
            return

        self.btn_freeze()
        self.info_text.set('正在生成Excel文档')
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()

        ws.append(self.title)
        for student in self.student_objs:
            ws.append(student.row)

        wb.save(path)
        wb.close()
        self.info_text.set('Excel文档已生成')
        self.btn_unfreeze()


if __name__ == "__main__":
    app = ttk.Window(title="成绩计算程序")
    app.iconbitmap('green_apple.ico')
    app.protocol('WM_DELETE_WINDOW', close_handle)  # 启用协议处理机制，点击关闭时按钮，触发事件
    app.place_window_center()
    App(app)
    app.mainloop()
