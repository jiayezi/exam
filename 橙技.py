"""
图像界面程序，辅助处理试题结构和小分表的信息，节约时间
版本：2.0
"""
import os
import threading
from tkinter import filedialog  # 文件访问对话框

import openpyxl
from ttkbootstrap.dialogs import Querybox, Messagebox
import ttkbootstrap as ttk
from ttkbootstrap.toast import ToastNotification
from openpyxl import Workbook, load_workbook
from win32com import client


def thread_it(func, *args):
    """将函数打包进线程"""
    # 创建
    t = threading.Thread(target=func, args=args)
    # 守护
    t.daemon = True
    # 启动
    t.start()


def unfreeze():
    """释放文本框，清空文本框"""
    info_text.config(state='normal')
    output_text.delete(1.0, 'end')  # 删除文本框里的内容
    info_text.delete(1.0, 'end')


def question_info():
    """提取从考试系统下载的整卷详情表里的题目信息。"""
    # 打开Excel表格
    file_path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                           defaultextension='.xlsx')
    if not file_path:
        return
    unfreeze()
    wb = load_workbook(file_path)

    text = '题号\t分数\t排序号\t是否主观题[0-否，1-是]\t客观题答案\n'
    counter = 0

    ws = wb['客观题']
    for i, row in enumerate(ws.values):
        if i < 3:
            continue
        if row[0] is None:
            break
        counter += 100
        text += f'{row[2]}\t{row[5]}\t{counter}\t0\t{row[4]}\n'
    ws = wb['主观题']
    for i, row in enumerate(ws.values):
        if i < 3:
            continue
        if row[0] is None:
            break
        counter += 100
        text += f'{row[1]}\t{row[2]}\t{counter}\t1\t\n'

    wb.close()
    output_text.insert('end', text)
    info_text.insert('end', '提取完成\n', 'center')
    output_text.focus()
    freeze()


def difficulty_level():
    """把数字放大100倍"""
    unfreeze()
    counter = 0
    data = input_text.get(1.0, 'end').strip()

    if data:
        data_list = data.split('\n')
        text = ''
        for i, s in enumerate(data_list):
            try:
                num = float(s)*100
            except ValueError:
                text += '\n'
                info_text.insert('end', f'第 {i + 1} 行不是纯数字，处理失败\n', 'center')
            else:
                text += f'{round(num)}\n'
                counter += 1

        output_text.insert('end', text.strip())
        info_text.insert('end', f'处理了 {counter} 个难度值\n', 'center')
        output_text.focus()
    freeze()


def single_choice():
    """提取字符串里的A、B、C、D、E、F、G、T"""
    unfreeze()
    data = input_text.get(1.0, 'end')  # 获取文本框里的数据
    if data:
        counter = 0
        text = ''
        for s in data:
            if s in ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'T'):
                text += f'{s}\n'
                counter += 1
        output_text.insert('end', text.strip())
        info_text.insert('end', f'提取了 {counter} 个答案\n', 'center')
        output_text.focus()
    freeze()


def skill_requirements():
    """把符号替换成该列对应的文字"""
    data = input_text.get(1.0, 'end').strip()
    if not data:
        return
    unfreeze()
    data_list = data.split('\n')
    title = data_list[0]
    content = data_list[1:]
    title_list = title.split('\t')

    text = ''
    for row_index, row in enumerate(content):
        row_list = row.split('\t')
        counter = 0
        for i, mark in enumerate(row_list):
            if mark.strip():
                counter += 1
                text += f'{title_list[i]}/'
        if counter == 0:
            text += f'\n'
            info_text.insert('end', f'第 {row_index + 2} 行没有符号\n', 'center')
        text = text[:-1]+'\n'
    output_text.insert('end', text.strip())

    info_text.insert('end', '全部处理完成\n', 'center')
    output_text.focus()
    freeze()


def OMR():
    """把制表符替换成逗号"""
    unfreeze()

    data = input_text.get(1.0, 'end').strip()
    # data_list = data.split('\n')
    # counter = 0
    # for row in data_list:
    #     row_list = row.split('\t')
        # for s in row_list:
        #     if len(s) == 1:
        #         text += f'{s},'
        #     else:
        #         text += '.,'
        #         counter += 1
        # text = text[:-1] + '\n'

    text = data.replace('\t', ',')
    output_text.insert('end', text)
    output_text.focus()
    freeze()


def multiple_OMR():
    """合并不定向选择答案"""
    unfreeze()

    data = input_text.get(1.0, 'end').strip()
    data_list = data.split('\n')

    for row in data_list:
        row_list = row.split('\t')
        for s in row_list:
            if len(s) > 0:
                output_text.insert('end', f'[{s}]')
            else:
                output_text.insert('end', '[.]')
        output_text.insert('end', '\n')

    output_text.focus()
    freeze()


def format_table():
    """格式化小分表，添加科目编号，添加班级列，检查选择题答案的数量，获取每个题的最大值"""
    # 打开Excel表格
    open_path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                           defaultextension='.xlsx')
    if not open_path:
        return
    unfreeze()
    excel = client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    wb = excel.Workbooks.Open(open_path, False)
    ws = wb.Worksheets(1)
    max_row = ws.UsedRange.Rows.Count

    # 从文件名提取编号
    file_name = os.path.split(open_path)[1]
    num = file_name[:file_name.rfind('.')]

    # 检查选择题答案的数量 找出这一列的最大值和最小值，如果两个值相等，就证明这一列的所有数字都一样
    range_obj = ws.Range('C2')
    range_obj.EntireColumn.Insert()
    ws.Cells(2, 3).Value = '=len(B2)'

    sourceRange = ws.Range(ws.Cells(2, 3), ws.Cells(2, 3))
    fillRange = ws.Range(ws.Cells(2, 3), ws.Cells(max_row, 3))
    sourceRange.AutoFill(Destination=fillRange)  # 模拟自动填充

    ws.Cells(max_row + 1, 3).Value = f'=MIN(C2:C{max_row})'
    ws.Cells(max_row + 2, 3).Value = f'=MAX(C2:C{max_row})'

    if ws.Cells(max_row + 1, 3).Value == ws.Cells(max_row + 2, 3).Value:
        info_text.insert('end', f'选择题答案数量一样\n', 'center')
    else:
        info_text.insert('end', '选择题答案数量不一样，请检查这个科目是否有多选题 Σ(ŎдŎ|||)ﾉﾉ\n', 'center')
    ws.Cells(max_row + 1, 3).Value = None
    ws.Cells(max_row + 2, 3).Value = None
    ws.Columns(3).Delete()

    # 找出每个小题的最大得分
    for col in range(3, ws.UsedRange.Columns.Count + 1):
        # 通过数字获取列号
        col_name_fun = f'=SUBSTITUTE(ADDRESS(1,{col},4),1,"")'
        ws.Cells(max_row + 1, col).Value = col_name_fun
        col_name = ws.Cells(max_row + 1, col).Value
        # 计算最大值
        ws.Cells(max_row + 2, col).Value = f'=MAX({col_name}2:{col_name}{max_row})'
        output_text.insert('end', f'{ws.Cells(max_row + 2, col).Value}\n')
    info_text.insert('end', '题目最高分查找完成\n', 'center')
    # 删除2行临时数据
    ws.Rows(max_row + 1).Delete()
    ws.Rows(max_row + 1).Delete()

    # 设置表格为文本格式
    ws.Cells.NumberFormatLocal = "@"

    # 插入单行单列
    range_obj = ws.Range('A1')
    range_obj.EntireRow.Insert()
    range_obj.EntireColumn.Insert()

    ws.Cells(1, 1).Value = num
    ws.Cells(2, 1).Value = '班级'
    ws.Cells(3, 1).Value = '0' * 16

    # 模拟自动填充
    sourceRange = ws.Range(ws.Cells(3, 1), ws.Cells(3, 1))
    fillRange = ws.Range(ws.Cells(3, 1), ws.Cells(ws.UsedRange.Rows.Count, 1))
    sourceRange.AutoFill(Destination=fillRange)

    wb.Close(SaveChanges=1)  # 保存并关闭
    excel.Quit()

    info_text.insert('end', '小分表修改完成\n', 'center')
    info_text.yview_moveto(1)
    output_text.focus()

    freeze()


def format_table_new():
    """格式化小分表，只修改小分表的标题和班级编号"""
    # 打开Excel表格
    open_path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                           defaultextension='.xlsx')
    if not open_path:
        return
    unfreeze()

    wb = openpyxl.load_workbook(open_path)
    ws = wb.active

    # 修改标题
    titles = ('班级', '考号', '卷面分', '折算分', '客观题答案')
    row1 = next(ws.rows)
    for i, value in enumerate(titles):
        row1[i].value = value
    for cell in row1:
        cell.value = cell.value.replace('_得分', '')

    # 选科的小分表的班级的前两位是科目编号
    subject_code = Querybox.get_string(prompt='请输入科目编号：', initialvalue='00')
    # max_values = []
    for col_index, col in enumerate(ws.columns):
        if col_index == 0:
            # 修改班级为6位编号
            for row_index, cell in enumerate(col):
                if row_index == 0:
                    continue
                class_str = ''
                cell_value = str(cell.value)
                for s in cell_value:
                    if s.isdigit():
                        class_str += s
                class_length = len(class_str)
                if class_length > 2:
                    class_str = class_str[-2:]
                elif class_length == 1:
                    class_str = '0'+class_str
                elif class_length == 0:
                    class_str = '99'
                cell.value = f'{subject_code}00{class_str}'
        # if col_index > 4:
        #     # 获取每一列的最大值
        #     max_value = 0
        #     for row_index, cell in enumerate(col):
        #         if row_index == 0:
        #             continue
        #         cell_value = cell.value
        #         if cell_value > max_value:
        #             max_value = cell_value
        #     max_values.append(max_value)
    wb.save(open_path)
    # for value in max_values:
    #     output_text.insert('end', f'{value}\n', 'center')

    info_text.insert('end', '小分表修改完成\n', 'center')
    info_text.yview_moveto(1)
    output_text.focus()
    freeze()


def format_table_om():
    """格式化从鸥玛考试系统下载的小分表"""
    # 打开Excel表格
    open_path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                           defaultextension='.xlsx')
    if not open_path:
        return
    unfreeze()

    wb = openpyxl.load_workbook(open_path)
    ws = wb.active
    wb2 = Workbook(write_only=True)
    ws2 = wb2.create_sheet()
    title = ['班级', '考号', '卷面分', '折算分', '客观题答案']

    # 选科的小分表的班级的前两位是科目编号
    subject_code = Querybox.get_string(prompt='请输入科目编号：', initialvalue='00')
    invalid_score = ('缺考', '缺扫', 0)
    score_start = 0
    for row_index, row in enumerate(ws.values):
        # 提取原始标题和题目起始索引
        if row_index == 2:
            for i, value in enumerate(row):
                if i > 12 and value[-2:] != '选项':
                    score_start = i
                    q_list = row[i: -1]
                    for q in q_list:
                        title.append(q.replace('_得分', ''))
                    ws2.append(title)
                    break
        if row_index < 3:
            continue

        # 提取班级、考号、科目分数
        score = row[8]
        if score in invalid_score:
            continue
        class_ = row[4]
        if len(class_) == 1:
            class_ = '0'+class_
        temp_list = [f'{subject_code}00{class_}', row[1], score, score]

        # 提取选择题答案
        text = ''
        for col_index in range(12, score_start):
            text += f'{row[col_index]},'
        temp_list.append(text[:-1])

        # 提取小分
        temp_list += row[score_start:-1]
        ws2.append(temp_list)

    wb.close()
    wb2.save(open_path)
    wb2.close()

    info_text.insert('end', '小分表修改完成\n', 'center')
    info_text.yview_moveto(1)
    output_text.focus()
    freeze()


def total_score_level():
    """把每个学生的单科成绩相加，计算总分"""
    all_data_list = []
    counter = 0

    def add_score_data():
        data = text0.get(1.0, 'end')
        if not data.strip():
            return
        unfreeze()
        nonlocal all_data_list
        nonlocal counter
        all_data_list.append(data)
        counter += 1
        ToastNotification(title='信息', message=f'已提交 {counter} 个科目成绩', duration=3000, position=(0, 220, 's'))\
            .show_toast()
        freeze()
        text0.delete(1.0, 'end')
        text0.focus()

    def calculate_total_score():
        nonlocal all_data_list
        if not all_data_list:
            return
        unfreeze()
        # 存储考号和分数的字典
        student_dict = {}
        titles = ['考号']
        for index, data in enumerate(all_data_list):
            data = data.strip()
            row_list = data.split('\n')

            try:
                for row in row_list:
                    singe_list = row.split('\t')
                    # 判断考号
                    student_id = singe_list[0]
                    if len(student_id) < 5 or not student_id.isdigit():
                        titles.append(singe_list[1])
                        continue
                    score = float(singe_list[1])

                    # 如果是新增的考号，为保证科目与分数对应，需要让该考号的之前的科目为0分
                    if student_id not in student_dict:
                        student_dict[student_id] = [0.0 for _ in range(index + 1)]
                    student_dict[student_id][index] = score
            except IndexError:
                info_text.insert('end', '数据不完整，处理失败，请同时提交考号和成绩 (ー_ー)!!\n', 'center')
                top.destroy()
                return
            except ValueError:
                info_text.insert('end', '成绩字段不是纯数字，处理失败 (ー_ー)!!\n', 'center')
                top.destroy()
                return
            # 把所有学生的下一个科目的分数初始化为0分
            if index < len(all_data_list) - 1:
                for key in student_dict:
                    student_dict[key].append(0.0)

        titles.append('总分')
        wb = Workbook()
        ws = wb.active
        ws.title = '总分表'

        # 添加首行
        if len(titles) > 2:
            ws.append(titles)
        else:
            row_data = ['考号'] + list(range(len(all_data_list))) + ['总分']
            ws.append(row_data)
        # 添加数据
        for row_index, key in enumerate(student_dict):
            row_data = [key]
            row_data.extend(student_dict[key])
            row_data.append(sum(student_dict[key]))
            ws.append(row_data)

        file_path = filedialog.asksaveasfilename(title='请选择文件存储路径',
                                                 initialdir='F:/用户目录/桌面/',
                                                 initialfile='总分表',
                                                 filetypes=[('Excel', '.xlsx')],
                                                 defaultextension='.xlsx')
        if file_path:
            wb.save(file_path)
            info_text.insert('end', '文件保存成功\n', 'center')
            wb.close()
        top.destroy()
        freeze()

    top = ttk.Toplevel()
    top.title('计算总分')
    top.place_window_center()

    lb = ttk.Label(top, text='考号和单科成绩：', font=('微软雅黑', 12))
    lb.pack(padx=20, pady=10)
    text0 = ttk.Text(top, width=100, height=25)
    text0.pack(padx=20)
    text0.focus()
    buttonbar = ttk.Frame(top, padding=20)
    buttonbar.pack()
    btn = ttk.Button(master=buttonbar, text='提交分数', command=add_score_data)
    btn.pack(side='left', padx=10, ipadx=8)
    btn = ttk.Button(master=buttonbar, text='计算总分', command=lambda: thread_it(calculate_total_score))
    btn.pack(side='left', padx=10, ipadx=8)

    top.mainloop()


def split_score_level():
    """按小题的分数拆分总分"""

    def paste_from_clipboard1(event):
        clipboard_text = root.clipboard_get()
        left_text.insert('end', clipboard_text)

    def paste_from_clipboard2(event):
        clipboard_text = root.clipboard_get()
        mid_text.insert('end', clipboard_text)

    def copy_to_clipboard(event):
        selected_text = right_text.get(1.0, 'end')
        root.clipboard_clear()
        root.clipboard_append(selected_text)

        info_text.config(state='normal')
        info_text.insert('end', '已复制到剪贴板\n', 'center')
        freeze()

    def split_score():
        data = left_text.get(1.0, 'end').strip()
        small_data = mid_text.get(1.0, 'end').strip()
        if not (data and small_data):
            return
        left_text.delete(1.0, 'end')
        mid_text.delete(1.0, 'end')
        right_text.delete(1.0, 'end')
        total_score_list = data.split('\n')
        small_score_list = small_data.split('\n')
        right_text.config(state='normal')
        text = ''
        try:
            for total_score in total_score_list:
                total_score = float(total_score)
                for small_score in small_score_list:
                    small_score = float(small_score)
                    if total_score > small_score:
                        text += f'{small_score}\t'
                        total_score -= small_score
                    else:
                        text += f'{total_score}\t'
                        total_score = 0
                text = text[:-1]+'\n'
        except ValueError:
            ToastNotification(title='信息', message='总分或题目分不是纯数字，拆分失败 (ー_ー)!!', duration=3000,
                              position=(0, 220, 's')).show_toast()
        else:
            right_text.insert('end', text[:-1])
            ToastNotification(title='信息', message='拆分完毕', duration=3000, position=(0, 220, 's')).show_toast()
            right_text.focus()

    top = ttk.Toplevel()
    top.title('拆分')
    top.iconbitmap('green_apple.ico')
    top.place_window_center()

    top.place_window_center()

    lb1 = ttk.Label(top, text='总分', font=('微软雅黑', 12))
    lb1.grid(row=0, column=0, padx=10, pady=10)
    lb2 = ttk.Label(top, text='小题满分', font=('微软雅黑', 12))
    lb2.grid(row=0, column=1, padx=10, pady=10)
    lb3 = ttk.Label(top, text='小分', font=('微软雅黑', 12))
    lb3.grid(row=0, column=2, padx=10, pady=10)

    left_text = ttk.Text(top, width=5, height=25)
    left_text.grid(row=1, column=0, padx=(20, 10))
    left_text.bind("<Double-Button-1>", paste_from_clipboard1)
    mid_text = ttk.Text(top, width=5, height=25)
    mid_text.grid(row=1, column=1, padx=10)
    mid_text.bind("<Double-Button-1>", paste_from_clipboard2)
    right_text = ttk.Text(top, width=90, height=25, state='disabled')
    right_text.grid(row=1, column=2, padx=(10, 20))
    right_text.bind("<Double-Button-1>", copy_to_clipboard)

    btn = ttk.Button(master=top, text='计算', command=split_score)
    btn.grid(row=2, column=0, ipadx=10, pady=20, columnspan=3)

    top.mainloop()


def freeze():
    """改变文本颜色，禁用文本框"""
    info_text.config(state='disabled')
    info_text.yview_moveto(1)  # 滚动到文本末尾
    input_text.delete(1.0, 'end')


def show_message():
    top = ttk.Toplevel()
    top.title('软件介绍')
    top.geometry(f'600x320')  # 窗口大小
    top.maxsize(700, 400)
    top.minsize(500, 200)
    top.place_window_center()

    text0 = ttk.Text(top, width=100, height=20, spacing2=10, spacing3=15)
    text0.pack()
    text0.insert('end', '题号：每行提取一个最多两位的数字\n'
                      '题目信息：提取从考试系统下载的整卷详情表里的题目信息\n'
                      '难度值：把数字放大100倍\n'
                      '单选答案：从文子里提取A-G的大写字母\n'
                      '能力要求：把文字里的标记符号替换成第一行对应的能力要求\n'
                      'OMR：合并所有列，把多选题答案和空白替换成“.”\n'
                      '多选OMR：合并所有列，把每个多选题答案放进中括号里\n'
                      '小分表：读取原始小分表，检查题目分数，整理小分表，另存为新的小分表\n'
                      '总分：输入考号和单科成绩，生成总分表\n'
                      '拆分：按照小题分数把每个学生的总分拆分成小分\n')

    text0.tag_add('forever', 1.0, 'end')
    text0.tag_config('forever', foreground='green', font=('黑体', 12))
    text0.config(state='disabled')
    top.mainloop()


def about():
    Messagebox.show_info(title='关于', message='橙技 2.0\n作者：迦叶子\n1601235906')


def paste_from_clipboard(event):
    clipboard_text = root.clipboard_get()
    input_text.insert('end', clipboard_text)


def copy_to_clipboard(event):
    selected_text = output_text.get(1.0, 'end')
    root.clipboard_clear()
    root.clipboard_append(selected_text)

    info_text.config(state='normal')
    info_text.insert('end', '已复制到剪贴板\n', 'center')
    freeze()


def close_handle():
    r = Messagebox.yesno(title='退出确认', message='确定要退出吗？')
    if r == '确认':
        root.destroy()


# 窗口
root = ttk.Window(themename='cerculean', title='橙技')
root.geometry(f'1180x820')  # 窗口大小
root.iconbitmap(bitmap='green_apple.ico')
root.iconbitmap(default='green_apple.ico')

# 菜单
main_menu = ttk.Menu(root)
sub_menu = ttk.Menu(main_menu, tearoff=0)
sub_menu.add_command(label='介绍', command=show_message)
sub_menu.add_command(label='关于', command=about)
main_menu.add_cascade(label='帮助', menu=sub_menu)
root.config(menu=main_menu)

root.protocol('WM_DELETE_WINDOW', close_handle)  # 启用协议处理机制，点击关闭时按钮，触发事件
root.place_window_center()

label1 = ttk.Label(root, text='原始数据', font=('黑体', 12))
label1.pack(pady=(20, 10))  # 按布局方式放置标签

input_text = ttk.Text(root, height=12)
input_text.pack(fill='x', padx=100)  # 文本框宽度沿水平方向自适应填充，左右两边空100像素
input_text.focus()
input_text.bind("<Double-Button-1>", paste_from_clipboard)

label2 = ttk.Label(root, text='计算结果', font=('黑体', 12))
label2.pack(pady=(20, 10))

output_text = ttk.Text(root, height=12)
output_text.pack(fill='x', padx=100)
# 为文本框绑定鼠标双击事件
output_text.bind("<Double-Button-1>", copy_to_clipboard)

info_text = ttk.Text(root, height=5, font=('黑体', 12), spacing3=8, border=-1, state='disabled')
info_text.pack(pady=10, padx=100, fill='x')
info_text.tag_config('center', foreground='green', justify='center')

# 按钮区域
buttonbar = ttk.Labelframe(root, text='选择功能', labelanchor='n', padding=20)
buttonbar.pack(pady=10,  padx=100)

btn = ttk.Button(master=buttonbar, text='题目信息', command=question_info)
btn.pack(side='left', padx=10)

btn = ttk.Button(master=buttonbar, text='难度值', command=difficulty_level)
btn.pack(side='left', padx=10)

btn = ttk.Button(master=buttonbar, text='单选答案', command=single_choice)
btn.pack(side='left', padx=10)

btn = ttk.Button(master=buttonbar, text='能力要求', command=skill_requirements)
btn.pack(side='left', padx=10)

btn = ttk.Button(master=buttonbar, text='OMR', command=OMR)
btn.pack(side='left', padx=10)

btn = ttk.Button(master=buttonbar, text='多选OMR', command=multiple_OMR)
btn.pack(side='left', padx=10)

btn = ttk.Button(master=buttonbar, text='小分表(鸥玛)', command=format_table_om)
btn.pack(side='left', padx=10)

btn = ttk.Button(master=buttonbar, text='小分表', command=format_table_new)
btn.pack(side='left', padx=10)

btn = ttk.Button(master=buttonbar, text='总分表', command=total_score_level)
btn.pack(side='left', padx=10)

btn = ttk.Button(master=buttonbar, text='拆分', command=split_score_level)
btn.pack(side='left', padx=10)

root.mainloop()
