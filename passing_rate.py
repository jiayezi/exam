"""找出每个科目的不及格的学生，存储到新的工作簿里"""
from openpyxl import load_workbook, Workbook
from tkinter import messagebox, filedialog  # 文件访问对话框


def student_rate():
    file_path = filedialog.askopenfilename(title='请选择Excel文件', filetypes=[('Excel', '.xlsx')],
                                           defaultextension='.xlsx')
    if file_path:
        # 新建工作薄
        wb2 = Workbook()

        # 创建每个科目的工作表
        for i, subject in enumerate(pas):
            wb2.create_sheet(subject)
            wb2[subject].append(titles)
        wb2.remove(wb2.worksheets[0])
        wb = load_workbook(file_path)
        ws = wb.active

        all_count, fail_count = 0, 0
        for col in range(extra+1,  ws.max_column+1):
            if ws.cell(1, col).value in pas:
                subject = ws.cell(1, col).value
            else:
                messagebox.showerror(message='没有找到科目，请选择正确格式的工作簿。')
                break
            score_index = col-1
            for row in ws.values:
                score = row[score_index]
                if isinstance(score, str):
                    continue
                if score > 0:
                    all_count += 1
                    if score < pas[subject]:
                        fail_count += 1

                        # 向新工作簿添加不及格的学生信息
                        fail_student = row[:extra]+(score,)
                        wb2[subject].append(fail_student)

            rate = fail_count / all_count
            rate_list = ['不及格率：', f'{rate:.1%}']
            wb2[subject].append(rate_list)
            rate_list = ['及格率：', f'{(1 - rate):.1%}']
            wb2[subject].append(rate_list)
            all_count, fail_count = 0, 0

        save_path = filedialog.asksaveasfilename(title='请选择文件存储路径',
                                                 initialdir='F:/用户目录/桌面/',
                                                 initialfile='玉林高一及格率、不及格率和人数汇总',
                                                 filetypes=[('Excel', '.xlsx')],
                                                 defaultextension='.xlsx')
        if save_path:
            wb2.save(save_path)


if __name__ == '__main__':
    pas = {'语文': 90, '数学': 90, '英语': 90, '政治': 60, '历史': 60, '地理': 60, '物理': 60, '化学': 60, '生物': 60}
    titles = ('学籍号', '考生号', '姓名', '学校', '班级', '成绩')
    extra = 5  # 前5列数据用不上
    student_rate()
