"""计算每个学生每个科目的赋分成绩、等级和排名
山东采用5等8级赋分制。
等级考试科目原始成绩从高到低划分为A、B+、B、C+、C、D+、D、E共8个等级。
参照正态分布原则，确定各等级人数所占比例分别为3%、7%、16%、24%、24%、16%、7%、3%。
等级考试科目成绩计入考生总成绩时，将A至E等级内的考生原始成绩，依照等比例转换法则，
分别转换到91-100、81-90、71-80、61-70、51-60、41-50、31-40、21-30八个分数区间，得到考生的等级成绩。
本程序把0分和空白的单元格视为缺考，不计算等级赋分"""

from tkinter import filedialog
from openpyxl import load_workbook, Workbook

extra = 14  # 前14列数据用不上

# 读取Excel文件
file_path = filedialog.askopenfilename(title='请选择Excel文件', initialdir='F:/用户目录/桌面/',
                                       filetypes=[('Excel', '.xlsx')], defaultextension='.xlsx')
if not file_path:
    exit()
wb = load_workbook(file_path, read_only=True)
ws = wb.active
ws_rows = []
for row in ws.values:
    ws_rows.append(list(row))
wb.close()
student_data = ws_rows[1:]
ws_title = ws_rows[0]
subjects = ws_title[extra:23]

rateT = (1, 0.97, 0.9, 0.74, 0.50, 0.26, 0.1, 0.07, 0)
rateY = ((100, 91), (90, 81), (80, 71), (70, 61), (60, 51), (50, 41), (40, 31), (30, 21))
dict_dj = {0: 'A', 1: 'B+', 2: 'B', 3: 'C+', 4: 'C', 5: 'D+', 6: 'D', 7: 'E'}


def sort_rule(score):
    if score is None or score == '':
        return 0
    else:
        return float(score)


for sub_index, subject in enumerate(subjects):
    student_data.sort(key=lambda x: sort_rule(x[extra + sub_index]), reverse=True)

    # 获取得分大于0分的人数、获取大于0分的最小原始分
    student_data_reverse = student_data[::-1]
    student_num = len(student_data)
    min_score = 0.0
    for w_index, row in enumerate(student_data_reverse):
        if row[extra + sub_index] is None or row[extra + sub_index] == '':
            continue
        if float(row[extra + sub_index]) > 0.0:
            student_num -= w_index
            min_score = float(row[extra + sub_index])
            break

    # 获取原始分等级区间
    rateS = [[float(student_data[0][extra + sub_index])]]
    temp_dj = 0
    rate = (student_num - 1) / student_num
    for row_index, row in enumerate(student_data):
        if row[extra + sub_index] is None or row[extra + sub_index] == '' or \
                student_data[row_index - 1][extra + sub_index] is None or \
                student_data[row_index - 1][extra + sub_index] == '':
            continue
        current_score = float(row[extra + sub_index])
        previous_score = float(student_data[row_index - 1][extra + sub_index])
        if current_score != previous_score:
            rate = (student_num - row_index - 1) / student_num  # 领先率
            for v_index, value in enumerate(rateT):
                if v_index == 0:
                    continue
                if rate >= value:
                    if temp_dj != v_index - 1:
                        temp_dj = v_index - 1
                        rateS[temp_dj - 1].append(float(student_data[row_index - 1][extra + sub_index]))
                        rateS.append([float(row[extra + sub_index])])
                    break
    # rateS[-1].append(float(student_data[-1][extra + i]))
    rateS[-1].append(min_score)

    print(f'\n{subject}原始分等级区间：{rateS}')

    # 计算赋分成绩和排名
    prev = -1  # 上个分数，初始值为-1
    rank = 0  # 当前排名
    for r_index, row in enumerate(student_data):
        score_str = row[extra + sub_index]
        if score_str is None or score_str == '' or float(score_str) < 0.001:
            row.append('')
            row.append('')
            row.append('')
            continue
        score = float(row[extra + sub_index])
        xsdj = 0
        for index, dj_score in enumerate(rateS):
            if index == 0:
                continue
            if dj_score[0] >= score >= dj_score[1]:
                xsdj = index
                break
        m = rateS[xsdj][1]
        n = rateS[xsdj][0]
        a = rateY[xsdj][1]
        b = rateY[xsdj][0]
        converts = (b * (score - m) + a * (n - score)) / (n - m)
        converts = round(converts)
        # print(f'等级：{xsdj}\tm：{m}\tn：{n}\ta：{a}\tb：{b}\t原始分：{score:.1f}\t转换分：{converts}')
        row.append(converts)
        row.append(dict_dj[xsdj])

        # 计算排名 如果分数不一样，排名就是索引值+1，如果分数一样，排名不变
        if converts != prev:
            rank = r_index + 1
            prev = converts
        row.append(rank)
    ws_title.append(f'{subject}转换分')
    ws_title.append(f'{subject}等级')
    ws_title.append(f'{subject}排名')

# 写入Excel文件
wb = Workbook(write_only=True)
ws = wb.create_sheet()
ws.append(ws_title)
for row in student_data:
    ws.append(row)
file_path = filedialog.asksaveasfilename(title='请选择文件存储路径', initialdir='F:/用户目录/桌面/',
                                         initialfile='赋分成绩',
                                         filetypes=[('Excel', '.xlsx')], defaultextension='.xlsx')
if file_path:
    wb.save(file_path)
    wb.close()

"""
  * @param m 原始分开始
  * @param n 原始分结束
  * @param a 等级赋值分开始
  * @param b 等级赋值分结束
  * @param score 实考分数
"""
