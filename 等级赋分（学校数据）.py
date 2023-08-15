"""按新高考的等级赋分制，计算每个学生每个科目的赋分成绩、等级和排名。
四川采用5等5级赋分制。
等级考试科目原始成绩从高到低划分为A、B、C、D、E共5个等级。
参照正态分布原则，确定各等级人数所占比例分别为15%、35%、35%、13%、2%。
等级考试科目成绩计入考生总成绩时，将A至E等级内的考生原始成绩，依照等比例转换法则，
分别转换到86-100、85-71、70-56、55-41、40-30五个分数区间，得到考生的等级成绩。
本程序把0分和空白的单元格视为缺考，不计算等级赋分"""

from tkinter import filedialog
from openpyxl import load_workbook, Workbook

extra = 12  # 忽略前几列的数据

# 读取Excel文件
file_path = filedialog.askopenfilename(title='请选择Excel文件', initialdir='F:/用户目录/桌面/',
                                       filetypes=[('Excel', '.xlsx')], defaultextension='.xlsx')
if not file_path:
    exit()
wb = load_workbook(file_path)
ws = wb.active
ws_rows = []
for row in ws.values:
    ws_rows.append(list(row))
student_data = ws_rows[1:]
ws_title = ws_rows[0]
subjects = ws_title[extra:]

# 配置领先率、赋分区间和等级
rateT = (1, 0.85, 0.5, 0.15, 0.02, 0)
rateY = ((100, 86), (85, 71), (70, 56), (55, 41), (40, 30))
dict_dj = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E'}


def sort_rule(score):
    """定义排序规则"""
    if isinstance(score, str) or score is None:
        return 0
    else:
        return score


def getp1(data, index, score):
    """获取小于某个数字的数量"""
    count = 0
    for row in data:
        current_str = row[index]
        if isinstance(current_str, str) or current_str is None:
            continue
        current_number = float(current_str)
        if score > current_number > 0:
            count += 1
    return count


for sub_index, subject in enumerate(subjects):
    student_data.sort(key=lambda x: sort_rule(x[extra + sub_index]), reverse=True)

    # 获取得分大于0分的人数、获取大于0分的最小原始分
    student_data_reverse = student_data[::-1]
    student_num = len(student_data)
    min_score = 0.0
    for w_index, row in enumerate(student_data_reverse):
        if isinstance(row[extra + sub_index], str) or row[extra + sub_index] is None:
            continue
        if float(row[extra + sub_index]) > 0.0:
            student_num -= w_index
            min_score = float(row[extra + sub_index])
            break

    # 获取原始分等级区间
    rateS = [[float(student_data[0][extra + sub_index])]]
    temp_dj = 0
    A = 0
    rate = (student_num - 1) / student_num
    for row_index, row in enumerate(student_data):
        current_score_str = row[extra + sub_index]
        if isinstance(current_score_str, str) or current_score_str is None:
            continue

        current_score = float(row[extra + sub_index])
        # 原始分为0分不参与原始分对照表
        if current_score < 0.001:
            continue

        A = getp1(student_data, extra + sub_index, current_score)
        rate = A / student_num  # 领先率

        for index, value in enumerate(rateT):
            if index == 0:
                continue
            if rate >= value:
                if temp_dj != index - 1:
                    temp_dj = index - 1
                    rateS[temp_dj - 1].append(float(student_data[row_index - 1][extra + sub_index]))
                    rateS.append([float(row[extra + sub_index])])
                break
        # print(temp_dj)

    rateS[-1].append(min_score)  # 转换分对照的原始分区间最后一个值为最后一个不为0分的最小值
    print(f'\n{subject}原始分等级区间：{rateS}')

    # 计算赋分成绩和排名
    prev = -1  # 上个分数，初始值为-1
    rank = 0  # 当前排名
    for r_index, row in enumerate(student_data):
        score = row[extra + sub_index]
        if not isinstance(score, (int, float)) or score == 0:
            row.append('')
            row.append('')
            row.append('')
            continue
        xsdj = -1
        converts = 0  # 转换分初始为0
        for index, dj_score in enumerate(rateS):
            if dj_score[0] >= score >= dj_score[1]:
                xsdj = index
                break

        m = rateS[xsdj][1]
        n = rateS[xsdj][0]
        a = rateY[xsdj][1]
        b = rateY[xsdj][0]
        if m == n:
            converts = (a+b)/2
        else:
            converts = (b * (score - m) + a * (n - score)) / (n - m)
        converts = round(converts)
        # print(f'等级：{xsdj}\tm：{m}\tn：{n}\ta：{a}\tb：{b}\t原始分：{score:.1f}\t转换分：{converts}')
        row.append(converts)
        row.append(dict_dj[xsdj])
        # 计算排名 如果分数不一样，排名就是索引值+1，如果分数一样，排名不变
        if converts != prev:
            rank = r_index + 1
            prev = converts
        row.append(rank)

    ws_title.append(f'{subject}转换分')
    ws_title.append(f'{subject}等级')
    ws_title.append(f'{subject}排名')
wb.close()

# 写入Excel文件
wb = Workbook()
ws = wb.active
ws.append(ws_title)
for row in student_data:
    ws.append(row)
file_path = filedialog.asksaveasfilename(title='请选择文件存储路径', initialdir='F:/用户目录/桌面/',
                                         initialfile='赋分成绩',
                                         filetypes=[('Excel', '.xlsx')], defaultextension='.xlsx')
if file_path:
    wb.save(file_path)
    wb.close()

"""
  * @param m 原始分开始
  * @param n 原始分结束
  * @param a 等级赋值分开始
  * @param b 等级赋值分结束
  * @param score 实考分数
"""
